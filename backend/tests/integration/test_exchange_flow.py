"""exchange 导入导出闭环 — 覆盖 student 与 honor 两阶段导入。"""
from __future__ import annotations

import io

from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit.models import AuditLog
from app.auth.models import Student, User, UserRole
from app.core.security import create_token
from app.exchange import repository as exchange_repo
from app.exchange.models import (
    BATCH_STATUS_VALIDATED,
    IMPORT_TYPE_TRANSCRIPT_PDF_REVIEW,
    StudentCourseRecord,
)
from app.exchange.service import ERROR_REPORT_COLUMN
from app.honor.models import HonorRecord


def _build_student_xlsx(rows: list[dict]) -> bytes:
    """按 _apply_student 期望的列头拼装一个最小 Excel。"""
    headers = [
        "student_no", "full_name", "gender", "birth_date",
        "grade_code", "major_code", "class_code",
        "political_status", "enrollment_year", "expected_graduation_year",
        "email", "status",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_honor_xlsx(rows: list[dict]) -> bytes:
    headers = [
        "category_code",
        "title",
        "level",
        "awarded_by",
        "document_no",
        "announced_at",
        "effective_from",
        "effective_to",
        "is_collective",
        "summary",
        "story_md",
        "acceptance_speech",
        "cover_image_url",
        "consent_flag",
        "student_no",
        "display_name",
        "major_snapshot",
        "grade_snapshot",
        "class_snapshot",
        "role_in_collective",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "honors"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _create_admin_headers(
    db: AsyncSession,
    *,
    work_no: str,
    display_name: str,
) -> dict[str, str]:
    user = User(work_no=work_no, display_name=display_name, is_active=True)
    db.add(user)
    await db.flush()
    db.add(UserRole(user_id=user.id, role_code="SUPER_ADMIN"))
    await db.commit()
    token = create_token(str(user.id), "access", extra_claims={"roles": ["SUPER_ADMIN"]})
    return {"Authorization": f"Bearer {token}"}


async def _latest_audit(
    db: AsyncSession,
    *,
    action: str,
) -> AuditLog | None:
    stmt = select(AuditLog).where(AuditLog.action == action).order_by(AuditLog.id.desc())
    return (await db.execute(stmt)).scalars().first()


async def test_student_import_happy_path_validate_and_commit(
    admin_client: AsyncClient, db: AsyncSession
) -> None:
    xlsx = _build_student_xlsx([
        {
            "student_no": "IMP2001",
            "full_name": "导入学生甲",
            "gender": "男",
            "grade_code": "2024",
            "major_code": "CS",
            "class_code": "CS2401",
            "enrollment_year": 2024,
            "expected_graduation_year": 2028,
        },
        {
            "student_no": "IMP2002",
            "full_name": "导入学生乙",
            "gender": "女",
            "grade_code": "2024",
            "major_code": "CS",
            "class_code": "CS2402",
        },
    ])

    upload = await admin_client.post(
        "/api/v1/admin/exchange/imports/student",
        files={
            "file": (
                "students.xlsx", xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 200, upload.text
    batch = upload.json()["data"]["batch"]
    assert batch["status"] == "VALIDATED"
    assert batch["total_rows"] == 2
    assert batch["fatal_rows"] == 0

    commit = await admin_client.post(
        f"/api/v1/admin/exchange/imports/{batch['id']}/commit",
        json={"note": "一期主档初始化"},
    )
    assert commit.status_code == 200, commit.text
    assert commit.json()["data"]["status"] == "COMMITTED"

    # 正式表落库：学生可查到
    stu = (
        await db.execute(select(Student).where(Student.student_no == "IMP2001"))
    ).scalar_one_or_none()
    assert stu is not None
    assert stu.full_name == "导入学生甲"


async def test_student_import_with_bad_row_fails_batch(
    admin_client: AsyncClient,
) -> None:
    xlsx = _build_student_xlsx([
        {"student_no": "IMP3001", "full_name": "正常行"},
        # 缺 student_no → FATAL
        {"student_no": None, "full_name": "缺学号"},
    ])
    upload = await admin_client.post(
        "/api/v1/admin/exchange/imports/student",
        files={"file": ("bad.xlsx", xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload.status_code == 200, upload.text
    batch = upload.json()["data"]["batch"]
    assert batch["status"] == "FAILED"
    assert batch["fatal_rows"] == 1
    assert batch["total_rows"] == 2

    # C-06：FAILED 批次不可 commit
    commit = await admin_client.post(
        f"/api/v1/admin/exchange/imports/{batch['id']}/commit",
        json={"note": "应被拒绝"},
    )
    assert commit.status_code == 400


async def test_v15_error_report_download_has_reason_column(
    admin_client: AsyncClient,
) -> None:
    """fix.md 修改点 4：失败批次返回带"错误原因"列的 Excel。"""
    xlsx = _build_student_xlsx([
        {"student_no": "IMP4001", "full_name": "正常行一"},
        {"student_no": None, "full_name": "缺学号两位"},   # FATAL
        {"student_no": "IMP4003", "full_name": None},      # FATAL（姓名空）
    ])
    upload = await admin_client.post(
        "/api/v1/admin/exchange/imports/student",
        files={"file": ("bad.xlsx", xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    batch_id = upload.json()["data"]["batch"]["id"]

    resp = await admin_client.get(
        f"/api/v1/admin/exchange/imports/{batch_id}/error-report"
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    wb = load_workbook(io.BytesIO(resp.content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # header 末列应是 错误原因
    header = list(rows[0])
    assert header[-1] == ERROR_REPORT_COLUMN
    # 至少有两条错误行，且 reason 非空
    err_rows = rows[1:]
    assert len(err_rows) == 2
    for r in err_rows:
        reason = r[-1]
        assert reason is not None and len(str(reason)) > 0
        # 应定位到行号与字段
        assert "row" in str(reason)


async def test_exchange_endpoints_reject_anonymous(client: AsyncClient) -> None:
    resp = await client.get("/api/v1/admin/exchange/imports")
    assert resp.status_code == 401


async def test_exchange_export_permissions_follow_policy_and_write_audit(
    client: AsyncClient,
    db: AsyncSession,
    counselor_headers: dict[str, str],
) -> None:
    db.add(
        Student(
            student_no="EXP-001",
            full_name="导出学生甲",
            grade_code="2024",
            major_code="CS",
            class_code="CS2401",
            political_status="共青团员",
            email="export@example.com",
        )
    )
    await db.commit()

    students_resp = await client.get(
        "/api/v1/admin/exchange/exports/students",
        headers=counselor_headers,
    )
    assert students_resp.status_code == 403, students_resp.text
    assert students_resp.json()["code"] == 40330

    transcripts_resp = await client.get(
        "/api/v1/admin/exchange/exports/transcripts",
        headers=counselor_headers,
    )
    assert transcripts_resp.status_code == 403, transcripts_resp.text
    assert transcripts_resp.json()["code"] == 40330

    curriculum_resp = await client.get(
        "/api/v1/admin/exchange/exports/curriculum",
        headers=counselor_headers,
    )
    assert curriculum_resp.status_code == 200, curriculum_resp.text
    assert curriculum_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    invalid_upload = await client.post(
        "/api/v1/admin/exchange/imports/student",
        headers=counselor_headers,
        files={
            "file": (
                "counselor-bad.xlsx",
                _build_student_xlsx([{"student_no": None, "full_name": "缺学号"}]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert invalid_upload.status_code == 200, invalid_upload.text
    batch_id = invalid_upload.json()["data"]["batch"]["id"]

    error_report_resp = await client.get(
        f"/api/v1/admin/exchange/imports/{batch_id}/error-report",
        headers=counselor_headers,
    )
    assert error_report_resp.status_code == 200, error_report_resp.text

    students_denied = await _latest_audit(db, action="EXPORT_STUDENTS_DENIED")
    assert students_denied is not None
    assert students_denied.result_code == "DENIED"

    transcripts_denied = await _latest_audit(db, action="EXPORT_TRANSCRIPTS_DENIED")
    assert transcripts_denied is not None
    assert transcripts_denied.result_code == "DENIED"

    curriculum_log = await _latest_audit(db, action="EXPORT_CURRICULUM")
    assert curriculum_log is not None
    assert curriculum_log.result_code == "SUCCESS"

    error_report_log = await _latest_audit(db, action="DOWNLOAD_ERROR_REPORT")
    assert error_report_log is not None
    assert error_report_log.result_code == "SUCCESS"


async def test_transcript_pdf_review_batch_cannot_be_committed_to_formal_records(
    admin_client: AsyncClient,
    db: AsyncSession,
) -> None:
    student = Student(
        student_no="PDF-REVIEW-001",
        full_name="PDF 核验学生",
        grade_code="2024",
        major_code="CS",
        class_code="CS2401",
    )
    db.add(student)
    await db.flush()
    batch = await exchange_repo.create_batch(
        db,
        batch_no="IM-TPDF-UNIT-001",
        import_type=IMPORT_TYPE_TRANSCRIPT_PDF_REVIEW,
        filename="transcript.pdf",
        file_size=128,
        mime_type="application/pdf",
        operator_id=None,
        operator_role="STUDENT",
    )
    await exchange_repo.finalize_batch(
        db,
        batch,
        status=BATCH_STATUS_VALIDATED,
        total_rows=1,
        ok_rows=0,
        warn_rows=1,
        fatal_rows=0,
        summary={
            "source": "STUDENT_TRANSCRIPT_PDF",
            "review_required": True,
            "formal_records_written": 0,
        },
    )
    await db.commit()

    commit = await admin_client.post(
        f"/api/v1/admin/exchange/imports/{batch.id}/commit",
        json={"note": "should not write"},
    )
    assert commit.status_code == 400, commit.text
    assert commit.json()["code"] == 40047
    assert "人工核验" in commit.json()["message"]

    records = (
        await db.execute(
            select(StudentCourseRecord).where(
                StudentCourseRecord.student_id == student.id
            )
        )
    ).scalars().all()
    assert records == []


async def test_honor_import_groups_rows_into_single_record(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    admin_headers = await _create_admin_headers(
        db,
        work_no="T400001",
        display_name="Import Honor Admin",
    )
    first_student = Student(
        student_no="IMP-H-001",
        full_name="荣誉导入甲",
        grade_code="2023",
        major_code="CS",
        class_code="CS2301",
    )
    second_student = Student(
        student_no="IMP-H-002",
        full_name="荣誉导入乙",
        grade_code="2023",
        major_code="CS",
        class_code="CS2302",
    )
    db.add_all([first_student, second_student])
    await db.commit()
    await db.refresh(first_student)
    await db.refresh(second_student)

    category = await client.post(
        "/api/v1/admin/honors/categories",
        headers=admin_headers,
        json={
            "code": "SCHOLARSHIP",
            "name": "奖学金",
            "description": "导入测试类目",
            "sort_order": 1,
            "is_active": True,
        },
    )
    assert category.status_code == 200, category.text

    xlsx = _build_honor_xlsx(
        [
            {
                "category_code": "SCHOLARSHIP",
                "title": "国家奖学金",
                "level": "NATIONAL",
                "awarded_by": "教育部",
                "document_no": "DOC-001",
                "announced_at": "2026-04-10",
                "effective_from": "2026-04-10",
                "effective_to": "2026-12-31",
                "is_collective": False,
                "summary": "同一荣誉导入分组",
                "story_md": "## 先进事迹",
                "acceptance_speech": "谢谢",
                "cover_image_url": "https://example.com/cover.png",
                "consent_flag": True,
                "student_no": "IMP-H-001",
            },
            {
                "category_code": "SCHOLARSHIP",
                "title": "国家奖学金",
                "level": "NATIONAL",
                "awarded_by": "教育部",
                "document_no": "DOC-001",
                "announced_at": "2026-04-10",
                "effective_from": "2026-04-10",
                "effective_to": "2026-12-31",
                "is_collective": False,
                "summary": "同一荣誉导入分组",
                "story_md": "## 先进事迹",
                "acceptance_speech": "谢谢",
                "cover_image_url": "https://example.com/cover.png",
                "consent_flag": True,
                "student_no": "IMP-H-002",
            },
        ]
    )

    upload = await client.post(
        "/api/v1/admin/exchange/imports/honor",
        headers=admin_headers,
        files={
            "file": (
                "honors.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 200, upload.text
    batch = upload.json()["data"]["batch"]
    assert batch["status"] == "VALIDATED"
    assert batch["fatal_rows"] == 0

    commit = await client.post(
        f"/api/v1/admin/exchange/imports/{batch['id']}/commit",
        headers=admin_headers,
        json={"note": "honor import"},
    )
    assert commit.status_code == 200, commit.text
    assert commit.json()["data"]["status"] == "COMMITTED"

    records = list((await db.execute(select(HonorRecord))).scalars().all())
    assert len(records) == 1
    record = records[0]
    assert record.title == "国家奖学金"
    assert len(record.recipients) == 2
    assert {recipient.student_no_snapshot for recipient in record.recipients} == {
        "IMP-H-001",
        "IMP-H-002",
    }
    assert {recipient.display_name for recipient in record.recipients} == {
        "荣誉导入甲",
        "荣誉导入乙",
    }


async def test_honor_import_error_report_and_unknown_student_commit_guard(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    admin_headers = await _create_admin_headers(
        db,
        work_no="T400002",
        display_name="Import Guard Admin",
    )
    category = await client.post(
        "/api/v1/admin/honors/categories",
        headers=admin_headers,
        json={
            "code": "MERIT",
            "name": "表彰",
            "description": "导入保护类目",
            "sort_order": 1,
            "is_active": True,
        },
    )
    assert category.status_code == 200, category.text

    invalid_xlsx = _build_honor_xlsx(
        [
            {
                "category_code": "MERIT",
                "title": "错误荣誉",
                "level": "INVALID",
                "awarded_by": "学院",
                "announced_at": "2026/04/10",
                "display_name": "未命中校验",
            }
        ]
    )
    invalid_upload = await client.post(
        "/api/v1/admin/exchange/imports/honor",
        headers=admin_headers,
        files={
            "file": (
                "honor-invalid.xlsx",
                invalid_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert invalid_upload.status_code == 200, invalid_upload.text
    invalid_batch_id = invalid_upload.json()["data"]["batch"]["id"]
    assert invalid_upload.json()["data"]["batch"]["status"] == "FAILED"

    error_report = await client.get(
        f"/api/v1/admin/exchange/imports/{invalid_batch_id}/error-report",
        headers=admin_headers,
    )
    assert error_report.status_code == 200, error_report.text
    report_wb = load_workbook(io.BytesIO(error_report.content), read_only=True)
    report_rows = list(report_wb.active.iter_rows(values_only=True))
    assert list(report_rows[0])[-1] == ERROR_REPORT_COLUMN
    assert "level" in str(report_rows[1][-1]) or "announced_at" in str(report_rows[1][-1])

    unknown_student_xlsx = _build_honor_xlsx(
        [
            {
                "category_code": "MERIT",
                "title": "待核验荣誉",
                "level": "SCHOOL",
                "awarded_by": "学院",
                "announced_at": "2026-04-10",
                "student_no": "NOT-FOUND-001",
            }
        ]
    )
    unknown_upload = await client.post(
        "/api/v1/admin/exchange/imports/honor",
        headers=admin_headers,
        files={
            "file": (
                "honor-unknown-student.xlsx",
                unknown_student_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert unknown_upload.status_code == 200, unknown_upload.text
    unknown_batch_id = unknown_upload.json()["data"]["batch"]["id"]
    assert unknown_upload.json()["data"]["batch"]["status"] == "VALIDATED"

    commit = await client.post(
        f"/api/v1/admin/exchange/imports/{unknown_batch_id}/commit",
        headers=admin_headers,
        json={"note": "should fail"},
    )
    assert commit.status_code == 400
    assert "NOT-FOUND-001" in commit.json()["message"]
