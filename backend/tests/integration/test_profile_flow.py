"""profile S3 集成测试：学籍只读、补录审批、scope 收口与快照导出。"""
from __future__ import annotations

import io
from datetime import UTC, datetime

from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit.models import AuditLog, RoleFieldPolicy
from app.audit.policies import POLITICAL_STATUS
from app.auth.models import ENROLLMENT_ACTIVE, ENROLLMENT_GRADUATED, Student, User, UserRole
from app.core.security import create_token
from app.profile.models import (
    PROFILE_APPROVAL_APPROVED,
    PROFILE_APPROVAL_PENDING,
    PROFILE_APPROVAL_REJECTED,
    PROFILE_FACT_COMPETITION,
    PROFILE_FACT_RESEARCH,
    PROFILE_FULL_VIEW_TARGET_PROFILE_FACTS,
    PROFILE_FULL_VIEW_TARGET_STUDENT_FIELD,
    PROFILE_SOURCE_STUDENT_SELF,
    PROFILE_SOURCE_TEACHER_ENTRY,
    ProfileCorrection,
    ProfileFact,
)


async def _create_student(
    db: AsyncSession,
    *,
    student_no: str,
    full_name: str,
    class_code: str,
    major_code: str,
    enrollment_status: str = ENROLLMENT_ACTIVE,
    enrollment_status_reason: str | None = None,
) -> Student:
    row = Student(
        student_no=student_no,
        full_name=full_name,
        grade_code="2023",
        major_code=major_code,
        class_code=class_code,
        enrollment_status=enrollment_status,
        enrollment_status_reason=enrollment_status_reason,
        enrollment_status_updated_at=datetime.now(UTC),
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return row


async def _create_headers(
    db: AsyncSession,
    *,
    work_no: str,
    display_name: str,
    student_id: int | None = None,
    roles: list[tuple[str, str | None]] | None = None,
) -> tuple[dict[str, str], User]:
    user = User(
        work_no=work_no,
        display_name=display_name,
        student_id=student_id,
        is_active=True,
    )
    db.add(user)
    await db.flush()
    role_codes: list[str] = []
    for role_code, scope_code in roles or []:
        db.add(UserRole(user_id=user.id, role_code=role_code, scope_code=scope_code))
        if role_code not in role_codes:
            role_codes.append(role_code)
    await db.commit()
    await db.refresh(user)
    claims: dict[str, object] = {"roles": role_codes}
    if student_id is not None:
        claims["sid"] = student_id
    token = create_token(str(user.id), "access", extra_claims=claims)
    return {"Authorization": f"Bearer {token}"}, user


async def _create_fact(
    db: AsyncSession,
    *,
    student_id: int,
    fact_type: str,
    title: str,
    source: str,
    approval_status: str,
    created_by: int | None,
    updated_by: int | None,
    is_sensitive: bool = False,
    review_comment: str | None = None,
) -> ProfileFact:
    extra = {"review_comment": review_comment} if review_comment else None
    row = ProfileFact(
        student_id=student_id,
        fact_type=fact_type,
        title=title,
        description=f"{title} description",
        source=source,
        approval_status=approval_status,
        is_sensitive=is_sensitive,
        created_by=created_by,
        updated_by=updated_by,
        approved_by=updated_by if approval_status == PROFILE_APPROVAL_APPROVED else None,
        approved_at=datetime.now(UTC) if approval_status == PROFILE_APPROVAL_APPROVED else None,
        extra=extra,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return row


async def _create_correction(
    db: AsyncSession,
    *,
    student_id: int,
    fact_id: int | None = None,
) -> ProfileCorrection:
    row = ProfileCorrection(
        student_id=student_id,
        fact_id=fact_id,
        field_name="description",
        current_value="old",
        proposed_value="new",
        reason="需要修正",
        status=PROFILE_APPROVAL_PENDING,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return row


async def _latest_audit(db: AsyncSession, *, action: str, entity_id: int) -> AuditLog | None:
    stmt = (
        select(AuditLog)
        .where(AuditLog.action == action, AuditLog.entity_id == entity_id)
        .order_by(AuditLog.id.desc())
    )
    return (await db.execute(stmt)).scalars().first()


async def _set_policy_mask(
    db: AsyncSession,
    *,
    role_code: str,
    field_name: str,
    mask_strategy: str,
) -> None:
    row = (
        await db.execute(
            select(RoleFieldPolicy).where(
                RoleFieldPolicy.role_code == role_code,
                RoleFieldPolicy.entity_code == POLITICAL_STATUS,
                RoleFieldPolicy.field_name == field_name,
            )
        )
    ).scalar_one()
    row.can_read = True
    row.mask_strategy = mask_strategy
    await db.commit()


async def test_profile_self_view_submission_review_and_metadata_visibility(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    student = await _create_student(
        db,
        student_no="P300001",
        full_name="画像学生甲",
        class_code="CS2301",
        major_code="CS",
        enrollment_status_reason="正常在读",
    )
    student_headers, student_user = await _create_headers(
        db,
        work_no="S500001",
        display_name="Profile Student",
        student_id=student.id,
        roles=[("STUDENT", None)],
    )
    admin_headers, admin_user = await _create_headers(
        db,
        work_no="T500001",
        display_name="Profile Admin",
        roles=[("SUPER_ADMIN", None)],
    )

    await _create_fact(
        db,
        student_id=student.id,
        fact_type=PROFILE_FACT_RESEARCH,
        title="科研项目 A",
        source=PROFILE_SOURCE_TEACHER_ENTRY,
        approval_status=PROFILE_APPROVAL_APPROVED,
        created_by=admin_user.id,
        updated_by=admin_user.id,
    )
    await _create_fact(
        db,
        student_id=student.id,
        fact_type=PROFILE_FACT_COMPETITION,
        title="敏感竞赛 B",
        source=PROFILE_SOURCE_TEACHER_ENTRY,
        approval_status=PROFILE_APPROVAL_APPROVED,
        created_by=admin_user.id,
        updated_by=admin_user.id,
        is_sensitive=True,
    )

    self_profile = await client.get("/api/v1/profile/me", headers=student_headers)
    assert self_profile.status_code == 200, self_profile.text
    self_data = self_profile.json()["data"]
    assert self_data["student"]["status"] == "IN_SCHOOL"
    assert self_data["student"]["enrollment_status"] == ENROLLMENT_ACTIVE
    assert self_data["student"]["enrollment_status_reason"] == "正常在读"
    assert [fact["title"] for fact in self_data["facts"]] == ["科研项目 A"]
    assert "created_by" not in self_data["facts"][0]
    assert "source_label" not in self_data["facts"][0]

    approve_submission = await client.post(
        "/api/v1/profile/me/facts",
        headers=student_headers,
        json={
            "fact_type": PROFILE_FACT_COMPETITION,
            "title": "学生补录通过项",
            "description": "待审批",
            "attachments": {
                "files": [
                    {
                        "name": "award-proof.pdf",
                        "path": "C:/fixtures/award-proof.pdf",
                        "size": 1024,
                        "mime_type": "application/pdf",
                    }
                ]
            },
        },
    )
    assert approve_submission.status_code == 200, approve_submission.text
    approved_fact_id = approve_submission.json()["data"]["id"]
    assert approve_submission.json()["data"]["approval_status"] == PROFILE_APPROVAL_PENDING
    assert approve_submission.json()["data"]["attachments"]["files"][0]["name"] == "award-proof.pdf"

    pending = await client.get(
        "/api/v1/admin/profile/facts/pending",
        headers=admin_headers,
    )
    assert pending.status_code == 200, pending.text
    pending_items = pending.json()["data"]["items"]
    assert len(pending_items) == 1
    assert pending_items[0]["id"] == approved_fact_id
    assert pending_items[0]["source_label"] == "学生补录"
    assert pending_items[0]["created_by_name"] == "Profile Student"
    assert pending_items[0]["updated_by_name"] == "Profile Student"

    approve = await client.post(
        f"/api/v1/admin/profile/facts/{approved_fact_id}/decision",
        headers=admin_headers,
        json={"decision": PROFILE_APPROVAL_APPROVED, "comment": "通过"},
    )
    assert approve.status_code == 200, approve.text
    assert approve.json()["data"]["approval_status"] == PROFILE_APPROVAL_APPROVED
    assert approve.json()["data"]["updated_by_name"] == "Profile Admin"
    assert approve.json()["data"]["review_comment"] == "通过"

    reject_submission = await client.post(
        "/api/v1/profile/me/facts",
        headers=student_headers,
        json={
            "fact_type": PROFILE_FACT_RESEARCH,
            "title": "学生补录驳回项",
            "description": "待驳回",
        },
    )
    assert reject_submission.status_code == 200, reject_submission.text
    rejected_fact_id = reject_submission.json()["data"]["id"]

    reject = await client.post(
        f"/api/v1/admin/profile/facts/{rejected_fact_id}/decision",
        headers=admin_headers,
        json={"decision": PROFILE_APPROVAL_REJECTED, "comment": "材料不足"},
    )
    assert reject.status_code == 200, reject.text
    assert reject.json()["data"]["approval_status"] == PROFILE_APPROVAL_REJECTED
    assert reject.json()["data"]["review_comment"] == "材料不足"

    submissions = await client.get(
        "/api/v1/profile/me/fact-submissions",
        headers=student_headers,
    )
    assert submissions.status_code == 200, submissions.text
    submissions_by_title = {
        item["title"]: item for item in submissions.json()["data"]["items"]
    }
    assert submissions_by_title["学生补录通过项"]["approval_status"] == PROFILE_APPROVAL_APPROVED
    assert submissions_by_title["学生补录通过项"]["review_comment"] == "通过"
    assert submissions_by_title["学生补录驳回项"]["approval_status"] == PROFILE_APPROVAL_REJECTED
    assert submissions_by_title["学生补录驳回项"]["review_comment"] == "材料不足"

    self_profile_after = await client.get("/api/v1/profile/me", headers=student_headers)
    assert self_profile_after.status_code == 200, self_profile_after.text
    visible_titles = [fact["title"] for fact in self_profile_after.json()["data"]["facts"]]
    assert "科研项目 A" in visible_titles
    assert "学生补录通过项" in visible_titles
    assert "学生补录驳回项" not in visible_titles
    assert "敏感竞赛 B" not in visible_titles

    admin_profile = await client.get(
        f"/api/v1/admin/profile/{student.id}",
        headers=admin_headers,
    )
    assert admin_profile.status_code == 200, admin_profile.text
    admin_facts = {item["title"]: item for item in admin_profile.json()["data"]["facts"]}
    assert admin_facts["学生补录通过项"]["source_label"] == "学生补录"
    assert admin_facts["学生补录通过项"]["created_by_name"] == "Profile Student"
    assert admin_facts["学生补录通过项"]["updated_by_name"] == "Profile Admin"
    assert admin_facts["学生补录通过项"]["review_comment"] == "通过"


async def test_profile_full_view_request_student_reveals_masked_fields_and_hidden_facts(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    student = await _create_student(
        db,
        student_no="P300021",
        full_name="完整查看学生",
        class_code="CS2301",
        major_code="CS",
    )
    student.political_status = "共青团员"
    await db.commit()
    await _set_policy_mask(
        db,
        role_code="STUDENT",
        field_name="political_status",
        mask_strategy="full",
    )
    student_headers, _ = await _create_headers(
        db,
        work_no="S500021",
        display_name="Full View Student",
        student_id=student.id,
        roles=[("STUDENT", None)],
    )
    admin_headers, admin_user = await _create_headers(
        db,
        work_no="T500021",
        display_name="Full View Admin",
        roles=[("SUPER_ADMIN", None)],
    )
    await _create_fact(
        db,
        student_id=student.id,
        fact_type=PROFILE_FACT_COMPETITION,
        title="隐藏竞赛记录",
        source=PROFILE_SOURCE_TEACHER_ENTRY,
        approval_status=PROFILE_APPROVAL_APPROVED,
        created_by=admin_user.id,
        updated_by=admin_user.id,
        is_sensitive=True,
    )

    masked_profile = await client.get("/api/v1/profile/me", headers=student_headers)
    assert masked_profile.status_code == 200, masked_profile.text
    masked_data = masked_profile.json()["data"]
    assert masked_data["student"]["political_status"] == "***"
    assert "political_status" in masked_data["masked_fields"]
    assert masked_data["hidden_sensitive_fact_count"] == 1
    assert [fact["title"] for fact in masked_data["facts"]] == []

    field_req = await client.post(
        "/api/v1/profile/me/full-view-requests",
        headers=student_headers,
        json={
            "target_type": PROFILE_FULL_VIEW_TARGET_STUDENT_FIELD,
            "field_name": "political_status",
            "reason": "核对个人材料",
        },
    )
    assert field_req.status_code == 200, field_req.text
    field_req_id = field_req.json()["data"]["id"]
    assert field_req.json()["data"]["status"] == PROFILE_APPROVAL_PENDING

    facts_req = await client.post(
        "/api/v1/profile/me/full-view-requests",
        headers=student_headers,
        json={
            "target_type": PROFILE_FULL_VIEW_TARGET_PROFILE_FACTS,
            "reason": "查看隐藏成长记录",
        },
    )
    assert facts_req.status_code == 200, facts_req.text
    facts_req_id = facts_req.json()["data"]["id"]

    my_requests = await client.get(
        "/api/v1/profile/me/full-view-requests",
        headers=student_headers,
    )
    assert my_requests.status_code == 200, my_requests.text
    assert {item["id"] for item in my_requests.json()["data"]["items"]} == {
        field_req_id,
        facts_req_id,
    }

    corrections = await client.get("/api/v1/profile/me/corrections", headers=student_headers)
    assert corrections.status_code == 200, corrections.text
    assert corrections.json()["data"]["items"] == []

    pending = await client.get(
        "/api/v1/admin/profile/full-view-requests",
        headers=admin_headers,
        params={"student_id": student.id, "status": PROFILE_APPROVAL_PENDING},
    )
    assert pending.status_code == 200, pending.text
    assert {item["id"] for item in pending.json()["data"]["items"]} == {
        field_req_id,
        facts_req_id,
    }

    for request_id in (field_req_id, facts_req_id):
        decision = await client.post(
            f"/api/v1/admin/profile/full-view-requests/{request_id}/decision",
            headers=admin_headers,
            json={"decision": PROFILE_APPROVAL_APPROVED, "comment": "同意查看"},
        )
        assert decision.status_code == 200, decision.text
        assert decision.json()["data"]["status"] == PROFILE_APPROVAL_APPROVED

    full_profile = await client.get("/api/v1/profile/me", headers=student_headers)
    assert full_profile.status_code == 200, full_profile.text
    full_data = full_profile.json()["data"]
    assert full_data["student"]["political_status"] == "共青团员"
    assert "political_status" not in full_data["masked_fields"]
    assert "political_status" in full_data["full_view_approved_fields"]
    assert full_data["full_view_sensitive_facts_approved"] is True
    assert full_data["hidden_sensitive_fact_count"] == 0
    assert [fact["title"] for fact in full_data["facts"]] == ["隐藏竞赛记录"]

    submit_log = await _latest_audit(
        db,
        action="FULL_VIEW_REQUEST_SUBMIT",
        entity_id=field_req_id,
    )
    assert submit_log is not None
    assert submit_log.entity_code == "PROFILE_FULL_VIEW_REQUEST"
    decide_log = await _latest_audit(
        db,
        action="FULL_VIEW_REQUEST_DECIDE",
        entity_id=field_req_id,
    )
    assert decide_log is not None
    assert decide_log.detail["refs"][0]["field_name"] == "political_status"


async def test_profile_full_view_request_counselor_reveals_masked_admin_field(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    student = await _create_student(
        db,
        student_no="P300022",
        full_name="辅导员完整查看学生",
        class_code="CS2301",
        major_code="CS",
    )
    student.political_status = "中共党员"
    await db.commit()
    await _set_policy_mask(
        db,
        role_code="COUNSELOR",
        field_name="political_status",
        mask_strategy="partial",
    )
    counselor_headers, counselor_user = await _create_headers(
        db,
        work_no="T500022",
        display_name="Scoped Counselor",
        roles=[("COUNSELOR", "CLASS:CS2301")],
    )
    admin_headers, _ = await _create_headers(
        db,
        work_no="T500023",
        display_name="Full View Super Admin",
        roles=[("SUPER_ADMIN", None)],
    )

    masked_profile = await client.get(
        f"/api/v1/admin/profile/{student.id}",
        headers=counselor_headers,
    )
    assert masked_profile.status_code == 200, masked_profile.text
    masked_data = masked_profile.json()["data"]
    assert masked_data["student"]["political_status"] != "中共党员"
    assert "political_status" in masked_data["masked_fields"]

    request_resp = await client.post(
        f"/api/v1/admin/profile/{student.id}/full-view-requests",
        headers=counselor_headers,
        json={
            "target_type": PROFILE_FULL_VIEW_TARGET_STUDENT_FIELD,
            "field_name": "political_status",
            "reason": "班级谈话前核对政治面貌",
        },
    )
    assert request_resp.status_code == 200, request_resp.text
    request_id = request_resp.json()["data"]["id"]
    assert request_resp.json()["data"]["requester_user_id"] == counselor_user.id

    approve = await client.post(
        f"/api/v1/admin/profile/full-view-requests/{request_id}/decision",
        headers=admin_headers,
        json={"decision": PROFILE_APPROVAL_APPROVED, "comment": "同意"},
    )
    assert approve.status_code == 200, approve.text

    full_profile = await client.get(
        f"/api/v1/admin/profile/{student.id}",
        headers=counselor_headers,
    )
    assert full_profile.status_code == 200, full_profile.text
    full_data = full_profile.json()["data"]
    assert full_data["student"]["political_status"] == "中共党员"
    assert "political_status" not in full_data["masked_fields"]
    assert full_data["full_view_approved_fields"] == ["political_status"]


async def test_profile_full_view_request_rejects_unsupported_field(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    student = await _create_student(
        db,
        student_no="P300099",
        full_name="非法字段学生",
        class_code="CS2301",
        major_code="CS",
    )
    student_headers, _ = await _create_headers(
        db,
        work_no="S509099",
        display_name="Unsupported Field Student",
        student_id=student.id,
        roles=[("STUDENT", None)],
    )

    resp = await client.post(
        "/api/v1/profile/me/full-view-requests",
        headers=student_headers,
        json={
            "target_type": PROFILE_FULL_VIEW_TARGET_STUDENT_FIELD,
            "field_name": "unsupported_field",
            "reason": "测试非法字段",
        },
    )
    assert resp.status_code == 400, resp.text
    assert resp.json()["code"] == 40190

    pending = await client.get(
        "/api/v1/profile/me/full-view-requests",
        headers=student_headers,
    )
    assert pending.status_code == 200, pending.text
    assert pending.json()["data"]["items"] == []


async def test_profile_scope_controls_search_detail_decisions_and_audit(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    await _create_student(
        db,
        student_no="P300002",
        full_name="班级命中学生",
        class_code="CS2301",
        major_code="CS",
    )
    await _create_student(
        db,
        student_no="P300003",
        full_name="旧 scope 学生",
        class_code="CS2302",
        major_code="CS",
    )
    out_scope_student = await _create_student(
        db,
        student_no="P300004",
        full_name="越权学生",
        class_code="EE2301",
        major_code="EE",
    )

    counselor_headers, _ = await _create_headers(
        db,
        work_no="T500002",
        display_name="Class Counselor",
        roles=[("COUNSELOR", "CLASS:CS2301")],
    )
    legacy_headers, _ = await _create_headers(
        db,
        work_no="T500003",
        display_name="Legacy Counselor",
        roles=[("COUNSELOR", "CS2302")],
    )
    major_headers, _ = await _create_headers(
        db,
        work_no="T500004",
        display_name="Major Teacher",
        roles=[("HEAD_TEACHER", "MAJOR:EE")],
    )

    counselor_search = await client.get(
        "/api/v1/admin/profile/students",
        headers=counselor_headers,
    )
    assert counselor_search.status_code == 200, counselor_search.text
    assert [item["student_no"] for item in counselor_search.json()["data"]["items"]] == ["P300002"]

    legacy_search = await client.get(
        "/api/v1/admin/profile/students",
        headers=legacy_headers,
    )
    assert legacy_search.status_code == 200, legacy_search.text
    assert [item["student_no"] for item in legacy_search.json()["data"]["items"]] == ["P300003"]

    major_search = await client.get(
        "/api/v1/admin/profile/students",
        headers=major_headers,
    )
    assert major_search.status_code == 200, major_search.text
    assert [item["student_no"] for item in major_search.json()["data"]["items"]] == ["P300004"]

    pending_fact = await _create_fact(
        db,
        student_id=out_scope_student.id,
        fact_type=PROFILE_FACT_RESEARCH,
        title="待审批越权补录",
        source=PROFILE_SOURCE_STUDENT_SELF,
        approval_status=PROFILE_APPROVAL_PENDING,
        created_by=None,
        updated_by=None,
    )
    correction = await _create_correction(
        db,
        student_id=out_scope_student.id,
        fact_id=pending_fact.id,
    )

    pending_list = await client.get(
        "/api/v1/admin/profile/facts/pending",
        headers=counselor_headers,
    )
    assert pending_list.status_code == 200, pending_list.text
    assert pending_list.json()["data"]["meta"]["total"] == 0

    detail_forbidden = await client.get(
        f"/api/v1/admin/profile/{out_scope_student.id}",
        headers=counselor_headers,
    )
    assert detail_forbidden.status_code == 403
    assert detail_forbidden.json()["code"] == 40321

    fact_decision_forbidden = await client.post(
        f"/api/v1/admin/profile/facts/{pending_fact.id}/decision",
        headers=counselor_headers,
        json={"decision": PROFILE_APPROVAL_APPROVED, "comment": "不应通过"},
    )
    assert fact_decision_forbidden.status_code == 403
    assert fact_decision_forbidden.json()["code"] == 40321

    correction_decision_forbidden = await client.post(
        f"/api/v1/admin/profile/corrections/{correction.id}/decision",
        headers=counselor_headers,
        json={
            "decision": PROFILE_APPROVAL_APPROVED,
            "comment": "不应处理",
            "apply_to_fact": True,
        },
    )
    assert correction_decision_forbidden.status_code == 403
    assert correction_decision_forbidden.json()["code"] == 40321

    snapshot_forbidden = await client.get(
        f"/api/v1/admin/profile/{out_scope_student.id}/snapshot.xlsx",
        headers=counselor_headers,
    )
    assert snapshot_forbidden.status_code == 403
    assert snapshot_forbidden.json()["code"] == 40321

    read_log = await _latest_audit(
        db,
        action="READ_ADMIN_DENIED",
        entity_id=out_scope_student.id,
    )
    assert read_log is not None
    assert read_log.result_code == "DENIED"

    fact_log = await _latest_audit(
        db,
        action="DECIDE_FACT_DENIED",
        entity_id=out_scope_student.id,
    )
    assert fact_log is not None
    assert fact_log.result_code == "DENIED"

    correction_log = await _latest_audit(
        db,
        action="DECIDE_CORRECTION_DENIED",
        entity_id=out_scope_student.id,
    )
    assert correction_log is not None
    assert correction_log.result_code == "DENIED"

    snapshot_log = await _latest_audit(
        db,
        action="EXPORT_SNAPSHOT_DENIED",
        entity_id=out_scope_student.id,
    )
    assert snapshot_log is not None
    assert snapshot_log.result_code == "DENIED"


async def test_profile_student_search_defaults_active_and_allows_explicit_history(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    await _create_student(
        db,
        student_no="P310001",
        full_name="默认在读学生",
        class_code="CS2401",
        major_code="CS",
    )
    await _create_student(
        db,
        student_no="P310002",
        full_name="毕业历史学生",
        class_code="CS2401",
        major_code="CS",
        enrollment_status=ENROLLMENT_GRADUATED,
    )
    admin_headers, _ = await _create_headers(
        db,
        work_no="T510001",
        display_name="Profile Search Admin",
        roles=[("SUPER_ADMIN", None)],
    )

    default_search = await client.get(
        "/api/v1/admin/profile/students",
        headers=admin_headers,
        params={"class_code": "CS2401"},
    )
    assert default_search.status_code == 200, default_search.text
    assert [item["student_no"] for item in default_search.json()["data"]["items"]] == [
        "P310001"
    ]

    include_history = await client.get(
        "/api/v1/admin/profile/students",
        headers=admin_headers,
        params={"class_code": "CS2401", "include_non_active": True},
    )
    assert include_history.status_code == 200, include_history.text
    assert {item["student_no"] for item in include_history.json()["data"]["items"]} == {
        "P310001",
        "P310002",
    }

    graduated_only = await client.get(
        "/api/v1/admin/profile/students",
        headers=admin_headers,
        params={"class_code": "CS2401", "enrollment_status": ENROLLMENT_GRADUATED},
    )
    assert graduated_only.status_code == 200, graduated_only.text
    assert [item["student_no"] for item in graduated_only.json()["data"]["items"]] == [
        "P310002"
    ]


async def test_profile_read_only_student_and_snapshot_exports(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    student = await _create_student(
        db,
        student_no="P300005",
        full_name="非在读学生",
        class_code="CS2305",
        major_code="CS",
        enrollment_status=ENROLLMENT_GRADUATED,
        enrollment_status_reason="已毕业",
    )
    student_headers, _ = await _create_headers(
        db,
        work_no="S500005",
        display_name="Readonly Student",
        student_id=student.id,
        roles=[("STUDENT", None)],
    )
    admin_headers, admin_user = await _create_headers(
        db,
        work_no="T500005",
        display_name="Snapshot Admin",
        roles=[("SUPER_ADMIN", None)],
    )
    await _create_fact(
        db,
        student_id=student.id,
        fact_type=PROFILE_FACT_RESEARCH,
        title="已归档科研记录",
        source=PROFILE_SOURCE_TEACHER_ENTRY,
        approval_status=PROFILE_APPROVAL_APPROVED,
        created_by=admin_user.id,
        updated_by=admin_user.id,
    )

    read_only_profile = await client.get("/api/v1/profile/me", headers=student_headers)
    assert read_only_profile.status_code == 200, read_only_profile.text
    assert read_only_profile.json()["data"]["student"]["enrollment_status"] == ENROLLMENT_GRADUATED
    assert read_only_profile.json()["data"]["student"]["enrollment_status_reason"] == "已毕业"

    correction_blocked = await client.post(
        "/api/v1/profile/me/corrections",
        headers=student_headers,
        json={
            "field_name": "description",
            "proposed_value": "申请修改",
            "reason": "测试只读限制",
        },
    )
    assert correction_blocked.status_code == 403
    assert correction_blocked.json()["code"] == 40311

    fact_blocked = await client.post(
        "/api/v1/profile/me/facts",
        headers=student_headers,
        json={
            "fact_type": PROFILE_FACT_RESEARCH,
            "title": "只读补录",
            "description": "不应允许",
        },
    )
    assert fact_blocked.status_code == 403
    assert fact_blocked.json()["code"] == 40311

    pdf_resp = await client.get(
        f"/api/v1/admin/profile/{student.id}/snapshot.pdf",
        headers=admin_headers,
    )
    assert pdf_resp.status_code == 200, pdf_resp.text
    assert pdf_resp.headers["content-type"].startswith("application/pdf")
    assert pdf_resp.content.startswith(b"%PDF")

    xlsx_resp = await client.get(
        f"/api/v1/admin/profile/{student.id}/snapshot.xlsx",
        headers=admin_headers,
    )
    assert xlsx_resp.status_code == 200, xlsx_resp.text
    assert xlsx_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(xlsx_resp.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0][:3] == ("student_no", "full_name", "fact_type")
    assert rows[1][0] == "P300005"
    assert rows[1][1] == "非在读学生"


async def test_admin_creates_student_updates_master_data_and_unbinds_wechat(
    client: AsyncClient,
    db: AsyncSession,
) -> None:
    admin_headers, _ = await _create_headers(
        db,
        work_no="T500006",
        display_name="Student Master Admin",
        roles=[("SUPER_ADMIN", None)],
    )

    created = await client.post(
        "/api/v1/admin/students",
        headers=admin_headers,
        json={
            "student_no": "P300006",
            "full_name": "新增学生",
            "gender": "女",
            "grade_code": "2024",
            "major_code": "CS",
            "class_code": "CS2406",
            "political_status": "共青团员",
            "enrollment_year": 2024,
            "expected_graduation_year": 2028,
        },
    )
    assert created.status_code == 200, created.text
    student = created.json()["data"]
    assert student["student_no"] == "P300006"
    assert student["full_name"] == "新增学生"

    updated = await client.patch(
        f"/api/v1/admin/students/{student['id']}/academic-info",
        headers=admin_headers,
        json={"full_name": "新增学生改名", "class_code": "CS2407"},
    )
    assert updated.status_code == 200, updated.text
    assert updated.json()["data"]["full_name"] == "新增学生改名"
    assert updated.json()["data"]["class_code"] == "CS2407"

    login = await client.post(
        "/api/v1/auth/wx-login",
        json={
            "code": "wx_profile_unbind",
            "student_no": "P300006",
            "full_name": "新增学生改名",
        },
    )
    assert login.status_code == 200, login.text

    binding = await client.get(
        f"/api/v1/admin/students/{student['id']}/wechat-binding",
        headers=admin_headers,
    )
    assert binding.status_code == 200, binding.text
    binding_data = binding.json()["data"]
    assert binding_data["bound"] is True
    assert binding_data["openid_masked"].startswith("mock_s")
    assert "STUDENT" in binding_data["roles"]

    unbound = await client.delete(
        f"/api/v1/admin/students/{student['id']}/wechat-binding",
        headers=admin_headers,
    )
    assert unbound.status_code == 200, unbound.text
    assert unbound.json()["data"]["bound"] is False

    user = (
        await db.execute(select(User).where(User.openid == "mock_student_P300006"))
    ).scalar_one()
    assert user.student_id is None
    roles = (
        await db.execute(select(UserRole.role_code).where(UserRole.user_id == user.id))
    ).scalars().all()
    assert "STUDENT" not in roles
    assert "GUEST" in roles


async def test_profile_new_endpoints_require_auth(client: AsyncClient) -> None:
    assert (await client.get("/api/v1/profile/me/fact-submissions")).status_code == 401
    assert (await client.get("/api/v1/profile/me/full-view-requests")).status_code == 401
    assert (
        await client.post(
            "/api/v1/profile/me/facts",
            json={"fact_type": PROFILE_FACT_RESEARCH, "title": "匿名补录"},
        )
    ).status_code == 401
    assert (
        await client.post(
            "/api/v1/profile/me/full-view-requests",
            json={"target_type": PROFILE_FULL_VIEW_TARGET_STUDENT_FIELD, "field_name": "political_status"},
        )
    ).status_code == 401
    assert (await client.get("/api/v1/admin/profile/facts/pending")).status_code == 401
    assert (await client.get("/api/v1/admin/profile/full-view-requests")).status_code == 401
    assert (await client.get("/api/v1/admin/profile/1/snapshot.pdf")).status_code == 401
