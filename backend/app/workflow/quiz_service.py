"""理论自测服务 — FR-005。

覆盖三种题型的判题、题库 CRUD、学生抽题与整卷提交。
"""
from __future__ import annotations

import csv
import io
import logging
import random
import uuid
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit.service import log_action
from app.core.exceptions import BizError, NotFoundError
from app.exchange import repository as exchange_repo
from app.exchange.models import (
    BATCH_STATUS_COMMITTED,
    BATCH_STATUS_FAILED,
    BATCH_STATUS_VALIDATED,
    ROW_RESULT_FAILED,
    ROW_RESULT_OK,
    ROW_SEVERITY_FATAL,
    ROW_SEVERITY_INFO,
    ROW_SEVERITY_WARN,
    ImportBatch,
)
from app.workflow.quiz_models import (
    QUIZ_DIFFICULTIES,
    QUIZ_TYPE_JUDGE,
    QUIZ_TYPE_MULTI,
    QUIZ_TYPE_SINGLE,
    QUIZ_TYPES,
    QuizQuestion,
    QuizRecord,
)

logger = logging.getLogger(__name__)

QUIZ_IMPORT_TYPE = "QUIZ_QUESTION"
QUIZ_IMPORT_TEMPLATE_COLUMNS = [
    "topic",
    "qtype",
    "stem",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_key",
    "explanation",
    "difficulty",
    "is_active",
    "source_name",
    "source_url",
]
QUIZ_IMPORT_MAX_BYTES = 10 * 1024 * 1024


# ======================================================================
# A. 判题
# ======================================================================
def _normalize_keys(raw: str) -> list[str]:
    """把 'a, c,D' 变成 ['A','C','D'] 且去重。"""
    parts = [p.strip().upper() for p in (raw or "").split(",")]
    return sorted({p for p in parts if p})


def _normalize_judge(raw: str) -> str:
    val = (raw or "").strip().upper()
    if val in {"TRUE", "T", "1", "YES", "Y", "对", "正确"}:
        return "TRUE"
    if val in {"FALSE", "F", "0", "NO", "N", "错", "错误"}:
        return "FALSE"
    return val


def _normalize_correct_key(qtype: str, correct_key: str) -> str:
    if qtype == QUIZ_TYPE_MULTI:
        return ",".join(_normalize_keys(correct_key))
    if qtype == QUIZ_TYPE_JUDGE:
        return _normalize_judge(correct_key)
    return (correct_key or "").strip().upper()


def grade_answer(question: QuizQuestion, answer: str) -> bool:
    """根据题型对答案判分。"""
    if question.qtype == QUIZ_TYPE_SINGLE:
        return (answer or "").strip().upper() == (question.correct_key or "").strip().upper()
    if question.qtype == QUIZ_TYPE_MULTI:
        return _normalize_keys(answer) == _normalize_keys(question.correct_key)
    if question.qtype == QUIZ_TYPE_JUDGE:
        return _normalize_judge(answer) == _normalize_judge(question.correct_key)
    raise BizError(f"未知题型：{question.qtype}")


# ======================================================================
# B. 校验入参
# ======================================================================
def _validate_question_payload(
    qtype: str,
    options_json: list | None,
    correct_key: str,
    difficulty: str | None,
) -> None:
    if qtype not in QUIZ_TYPES:
        raise BizError(f"题型必须是 {QUIZ_TYPES} 之一")
    if difficulty is not None and difficulty not in QUIZ_DIFFICULTIES:
        raise BizError(f"难度必须是 {QUIZ_DIFFICULTIES} 之一")

    if qtype in (QUIZ_TYPE_SINGLE, QUIZ_TYPE_MULTI):
        if not options_json or len(options_json) < 2:
            raise BizError("选择题至少需要 2 个选项")
        keys = {str(opt.get("key", "")).strip().upper() for opt in options_json}
        if len(keys) != len(options_json) or "" in keys:
            raise BizError("选项 key 必须非空且唯一")
        correct = _normalize_keys(correct_key)
        if not correct:
            raise BizError("correct_key 不能为空")
        if any(k not in keys for k in correct):
            raise BizError("correct_key 必须是 options 中的 key")
        if qtype == QUIZ_TYPE_SINGLE and len(correct) != 1:
            raise BizError("单选题只能有一个正确答案")
        if qtype == QUIZ_TYPE_MULTI and len(correct) < 2:
            raise BizError("多选题至少两个正确答案")
    elif qtype == QUIZ_TYPE_JUDGE:
        if _normalize_judge(correct_key) not in {"TRUE", "FALSE"}:
            raise BizError("判断题正确答案必须是 TRUE 或 FALSE")


# ======================================================================
# C. 题库 CRUD（管理端）
# ======================================================================
async def create_question(
    db: AsyncSession,
    *,
    topic: str,
    qtype: str,
    stem: str,
    options_json: list | None,
    correct_key: str,
    explanation: str | None,
    difficulty: str | None,
    source_name: str | None = None,
    source_url: str | None = None,
    source_official: bool = False,
    import_batch_id: int | None = None,
    operator_id: int,
    operator_role: str | None,
) -> QuizQuestion:
    _validate_question_payload(qtype, options_json, correct_key, difficulty)
    # 多选正则化后再落库，避免大小写/顺序歧义
    stored_correct = _normalize_correct_key(qtype, correct_key)
    row = QuizQuestion(
        topic=topic.strip(),
        qtype=qtype,
        stem=stem.strip(),
        options_json=options_json,
        correct_key=stored_correct,
        explanation=explanation,
        difficulty=difficulty,
        source_name=(source_name or "").strip() or None,
        source_url=(source_url or "").strip() or None,
        source_official=bool(source_official),
        import_batch_id=import_batch_id,
        created_by=operator_id,
    )
    db.add(row)
    await db.flush()
    await log_action(
        db,
        event_type="QUIZ",
        entity_code="QUIZ_QUESTION",
        entity_id=row.id,
        action="CREATE",
        actor_user_id=operator_id,
        actor_role=operator_role,
        detail={"topic": topic, "qtype": qtype},
    )
    await db.commit()
    await db.refresh(row)
    return row


async def update_question(
    db: AsyncSession,
    question_id: int,
    *,
    topic: str | None = None,
    qtype: str | None = None,
    stem: str | None = None,
    options_json: list | None = None,
    correct_key: str | None = None,
    explanation: str | None = None,
    difficulty: str | None = None,
    source_name: str | None = None,
    source_url: str | None = None,
    source_official: bool | None = None,
    is_active: bool | None = None,
    operator_id: int,
    operator_role: str | None,
) -> QuizQuestion:
    row = await db.get(QuizQuestion, question_id)
    if row is None:
        raise NotFoundError("题目不存在")

    new_qtype = qtype or row.qtype
    new_options = options_json if options_json is not None else row.options_json
    new_correct = correct_key if correct_key is not None else row.correct_key
    new_difficulty = difficulty if difficulty is not None else row.difficulty
    _validate_question_payload(new_qtype, new_options, new_correct, new_difficulty)

    if topic is not None:
        row.topic = topic.strip()
    if qtype is not None:
        row.qtype = qtype
    if stem is not None:
        row.stem = stem.strip()
    if options_json is not None:
        row.options_json = options_json
    if correct_key is not None:
        if new_qtype == QUIZ_TYPE_MULTI:
            row.correct_key = ",".join(_normalize_keys(correct_key))
        elif new_qtype == QUIZ_TYPE_JUDGE:
            row.correct_key = _normalize_judge(correct_key)
        else:
            row.correct_key = correct_key.strip().upper()
    if explanation is not None:
        row.explanation = explanation
    if difficulty is not None:
        row.difficulty = difficulty
    if source_name is not None:
        row.source_name = source_name.strip() or None
    if source_url is not None:
        row.source_url = source_url.strip() or None
    if source_official is not None:
        row.source_official = bool(source_official)
    if is_active is not None:
        row.is_active = is_active

    await db.flush()
    await log_action(
        db,
        event_type="QUIZ",
        entity_code="QUIZ_QUESTION",
        entity_id=row.id,
        action="UPDATE",
        actor_user_id=operator_id,
        actor_role=operator_role,
    )
    await db.commit()
    await db.refresh(row)
    return row


def _gen_import_batch_no() -> str:
    return (
        f"IM-QUIZ-{datetime.now(UTC).strftime('%y%m%d%H%M%S')}-"
        f"{uuid.uuid4().hex[:6].upper()}"
    )


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _parse_bool(raw: Any, *, default: bool = True) -> bool | None:
    value = _clean(raw).lower()
    if value == "":
        return default
    if value in {"1", "true", "yes", "y", "启用", "是", "active"}:
        return True
    if value in {"0", "false", "no", "n", "停用", "否", "inactive"}:
        return False
    return None


def _read_csv_rows(file_bytes: bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk")
    reader = csv.DictReader(io.StringIO(text))
    header = [_clean(item) for item in (reader.fieldnames or [])]
    rows = []
    for row_no, row in enumerate(reader, start=2):
        item = {_clean(key): _clean(value) for key, value in row.items() if key is not None}
        if any(item.values()):
            rows.append((row_no, item))
    return header, rows


def _read_xlsx_rows(file_bytes: bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    try:
        header = [_clean(item) for item in next(row_iter)]
    except StopIteration:
        return [], []
    rows = []
    for row_no, raw in enumerate(row_iter, start=2):
        values = [_clean(item) for item in raw]
        if not any(values):
            continue
        rows.append(
            (
                row_no,
                {
                    header[index]: values[index]
                    for index in range(min(len(header), len(values)))
                    if header[index]
                },
            )
        )
    return header, rows


def _read_import_rows(filename: str, file_bytes: bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return _read_csv_rows(file_bytes)
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        return _read_xlsx_rows(file_bytes)
    raise BizError("理论自测题库仅支持 .xlsx / .csv 文件", code=40088)


def _validate_import_header(header: list[str]) -> str | None:
    normalized = [item.strip() for item in header if item.strip()]
    if normalized != QUIZ_IMPORT_TEMPLATE_COLUMNS:
        return f"模板列必须为：{', '.join(QUIZ_IMPORT_TEMPLATE_COLUMNS)}"
    return None


def _normalize_import_row(raw: dict[str, Any]) -> tuple[str, str | None, str | None, dict[str, Any] | None]:
    topic = _clean(raw.get("topic"))
    qtype = _clean(raw.get("qtype")).upper()
    stem = _clean(raw.get("stem"))
    correct_key = _clean(raw.get("correct_key"))
    difficulty = _clean(raw.get("difficulty")).upper() or None
    is_active = _parse_bool(raw.get("is_active"), default=True)
    source_name = _clean(raw.get("source_name")) or None
    source_url = _clean(raw.get("source_url")) or None
    source_official = bool(source_url)

    if not topic:
        return ROW_SEVERITY_FATAL, "topic", "主题不能为空", None
    if not stem:
        return ROW_SEVERITY_FATAL, "stem", "题干不能为空", None
    if qtype not in QUIZ_TYPES:
        return ROW_SEVERITY_FATAL, "qtype", f"题型必须是 {QUIZ_TYPES} 之一", None
    if not correct_key:
        return ROW_SEVERITY_FATAL, "correct_key", "正确答案不能为空", None
    if is_active is None:
        return ROW_SEVERITY_FATAL, "is_active", "is_active 仅支持 true/false/启用/停用", None

    options_json: list[dict[str, str]] | None = None
    if qtype in (QUIZ_TYPE_SINGLE, QUIZ_TYPE_MULTI):
        options_json = []
        for key in ("A", "B", "C", "D"):
            text = _clean(raw.get(f"option_{key.lower()}"))
            if text:
                options_json.append({"key": key, "text": text})

    try:
        _validate_question_payload(qtype, options_json, correct_key, difficulty)
    except BizError as exc:
        return ROW_SEVERITY_FATAL, "question", exc.message, None

    normalized = {
        "topic": topic,
        "qtype": qtype,
        "stem": stem,
        "options_json": options_json,
        "correct_key": _normalize_correct_key(qtype, correct_key),
        "explanation": _clean(raw.get("explanation")) or None,
        "difficulty": difficulty,
        "is_active": bool(is_active),
        "source_name": source_name,
        "source_url": source_url,
        "source_official": source_official,
    }
    return ROW_SEVERITY_INFO, None, None, normalized


async def _find_question_by_topic_stem(
    db: AsyncSession,
    *,
    topic: str,
    stem: str,
) -> QuizQuestion | None:
    stmt = select(QuizQuestion).where(QuizQuestion.topic == topic, QuizQuestion.stem == stem)
    return (await db.execute(stmt)).scalar_one_or_none()


async def preview_question_import(
    db: AsyncSession,
    *,
    filename: str,
    file_bytes: bytes,
    mime_type: str | None,
    operator_id: int,
    operator_role: str | None,
) -> ImportBatch:
    if not file_bytes:
        raise BizError("题库导入文件为空", code=40089)
    if len(file_bytes) > QUIZ_IMPORT_MAX_BYTES:
        raise BizError("题库导入文件超过 10MB 上限", code=41311, http_status=413)

    batch = await exchange_repo.create_batch(
        db,
        batch_no=_gen_import_batch_no(),
        import_type=QUIZ_IMPORT_TYPE,
        filename=filename,
        file_size=len(file_bytes),
        mime_type=mime_type,
        operator_id=operator_id,
        operator_role=operator_role,
    )

    fatal = warn = ok_rows = total = 0
    seen: set[tuple[str, str]] = set()
    try:
        header, upload_rows = _read_import_rows(filename, file_bytes)
    except Exception as exc:
        await exchange_repo.add_batch_row(
            db,
            batch_id=batch.id,
            row_no=1,
            severity=ROW_SEVERITY_FATAL,
            result=ROW_RESULT_FAILED,
            field_name="file",
            message=f"文件解析失败：{exc}"[:512],
        )
        await exchange_repo.finalize_batch(
            db,
            batch,
            status=BATCH_STATUS_FAILED,
            total_rows=0,
            ok_rows=0,
            warn_rows=0,
            fatal_rows=1,
            summary={"template_columns": QUIZ_IMPORT_TEMPLATE_COLUMNS, "error": str(exc)[:256]},
        )
        await log_action(
            db,
            event_type="QUIZ",
            entity_code="QUIZ_IMPORT_BATCH",
            action="PREVIEW_IMPORT",
            entity_id=batch.id,
            actor_user_id=operator_id,
            actor_role=operator_role,
            result_code="FAILED",
            detail={"filename": filename, "fatal": 1},
        )
        await db.commit()
        await db.refresh(batch)
        return batch

    header_error = _validate_import_header(header)
    if header_error:
        fatal = 1
        await exchange_repo.add_batch_row(
            db,
            batch_id=batch.id,
            row_no=1,
            severity=ROW_SEVERITY_FATAL,
            result=ROW_RESULT_FAILED,
            field_name="header",
            message=header_error,
            raw_data={"header": header},
        )
    else:
        for row_no, raw in upload_rows:
            total += 1
            severity, field, message, normalized = _normalize_import_row(raw)
            if normalized is not None:
                dedupe_key = (normalized["topic"], normalized["stem"])
                if dedupe_key in seen:
                    severity, field, message, normalized = (
                        ROW_SEVERITY_FATAL,
                        "stem",
                        "同一文件内存在重复的 topic + stem",
                        None,
                    )
                else:
                    seen.add(dedupe_key)
                    existing = await _find_question_by_topic_stem(
                        db,
                        topic=dedupe_key[0],
                        stem=dedupe_key[1],
                    )
                    if existing is not None:
                        severity = ROW_SEVERITY_WARN
                        message = "提交时将更新已有题目"
            if severity == ROW_SEVERITY_FATAL:
                fatal += 1
                result = ROW_RESULT_FAILED
            elif severity == ROW_SEVERITY_WARN:
                warn += 1
                ok_rows += 1
                result = ROW_RESULT_OK
            else:
                ok_rows += 1
                result = ROW_RESULT_OK
            await exchange_repo.add_batch_row(
                db,
                batch_id=batch.id,
                row_no=row_no,
                severity=severity,
                result=result,
                field_name=field,
                message=message,
                raw_data={"raw": raw, "normalized": normalized},
            )

    status = BATCH_STATUS_VALIDATED if fatal == 0 else BATCH_STATUS_FAILED
    await exchange_repo.finalize_batch(
        db,
        batch,
        status=status,
        total_rows=total,
        ok_rows=ok_rows,
        warn_rows=warn,
        fatal_rows=fatal,
        summary={"template_columns": QUIZ_IMPORT_TEMPLATE_COLUMNS},
    )
    await log_action(
        db,
        event_type="QUIZ",
        entity_code="QUIZ_IMPORT_BATCH",
        action="PREVIEW_IMPORT",
        entity_id=batch.id,
        actor_user_id=operator_id,
        actor_role=operator_role,
        result_code="FAILED" if fatal else "SUCCESS",
        detail={
            "filename": filename,
            "total": total,
            "ok": ok_rows,
            "warn": warn,
            "fatal": fatal,
        },
    )
    await db.commit()
    await db.refresh(batch)
    return batch


async def commit_question_import(
    db: AsyncSession,
    batch_id: int,
    *,
    operator_id: int,
    operator_role: str | None,
) -> tuple[ImportBatch, dict[str, int]]:
    batch = await exchange_repo.get_batch(db, batch_id)
    if batch is None:
        raise NotFoundError("题库导入批次不存在")
    if batch.import_type != QUIZ_IMPORT_TYPE:
        raise BizError("该批次不是理论自测题库导入", code=40090)
    if batch.status == BATCH_STATUS_COMMITTED:
        raise BizError("该题库导入批次已提交", code=40091)
    if batch.status != BATCH_STATUS_VALIDATED or batch.fatal_rows > 0:
        raise BizError("仅 VALIDATED 且无致命错误的题库导入批次可提交", code=40092)

    rows = await exchange_repo.list_batch_rows(db, batch_id, limit=100000)
    created = updated = skipped = 0
    for row in rows:
        if row.severity == ROW_SEVERITY_FATAL:
            skipped += 1
            continue
        normalized = (row.raw_data or {}).get("normalized")
        if not isinstance(normalized, dict):
            skipped += 1
            continue
        existing = await _find_question_by_topic_stem(
            db,
            topic=normalized["topic"],
            stem=normalized["stem"],
        )
        if existing is None:
            db.add(
                QuizQuestion(
                    **normalized,
                    import_batch_id=batch.id,
                    created_by=operator_id,
                )
            )
            created += 1
            continue
        for field in (
            "qtype",
            "options_json",
            "correct_key",
            "explanation",
            "difficulty",
            "is_active",
            "source_name",
            "source_url",
            "source_official",
        ):
            setattr(existing, field, normalized.get(field))
        existing.import_batch_id = batch.id
        updated += 1

    await exchange_repo.mark_batch_committed(db, batch)
    batch.summary = {
        **(batch.summary or {}),
        "commit": {
            "created": created,
            "updated": updated,
            "skipped": skipped,
        },
    }
    await log_action(
        db,
        event_type="QUIZ",
        entity_code="QUIZ_IMPORT_BATCH",
        action="COMMIT_IMPORT",
        entity_id=batch.id,
        actor_user_id=operator_id,
        actor_role=operator_role,
        detail={
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "batch_no": batch.batch_no,
        },
    )
    await db.commit()
    await db.refresh(batch)
    return batch, {"created": created, "updated": updated, "skipped": skipped}


def build_import_template(fmt: str = "xlsx") -> tuple[bytes, str, str]:
    normalized = (fmt or "xlsx").lower()
    sample = [
        "示例主题",
        "SINGLE",
        "示例题干：请替换为有来源的官方题目",
        "示例选项A",
        "示例选项B",
        "示例选项C",
        "示例选项D",
        "A",
        "示例解析：请填写可追溯来源或留空。",
        "EASY",
        "true",
        "共产党员网知识自测",
        "https://www.12371.cn/special/zszc/",
    ]
    if normalized == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(QUIZ_IMPORT_TEMPLATE_COLUMNS)
        writer.writerow(sample)
        return (
            ("\ufeff" + buffer.getvalue()).encode("utf-8"),
            "quiz-question-import-template.csv",
            "text/csv; charset=utf-8",
        )
    if normalized != "xlsx":
        raise BizError("format 仅支持 xlsx/csv", code=40093)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "quiz_questions"
    sheet.append(QUIZ_IMPORT_TEMPLATE_COLUMNS)
    sheet.append(sample)
    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "quiz-question-import-template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


async def delete_question(
    db: AsyncSession,
    question_id: int,
    *,
    operator_id: int,
    operator_role: str | None,
) -> None:
    """软删除：把 is_active 置 False，保留历史作答记录的外键完整性。"""
    row = await db.get(QuizQuestion, question_id)
    if row is None:
        raise NotFoundError("题目不存在")
    row.is_active = False
    await db.flush()
    await log_action(
        db,
        event_type="QUIZ",
        entity_code="QUIZ_QUESTION",
        entity_id=row.id,
        action="DELETE",
        actor_user_id=operator_id,
        actor_role=operator_role,
    )
    await db.commit()


async def list_questions(
    db: AsyncSession,
    *,
    topic: str | None = None,
    qtype: str | None = None,
    is_active: bool | None = None,
    keyword: str | None = None,
    page: int = 1,
    size: int = 20,
) -> tuple[list[QuizQuestion], int]:
    stmt = select(QuizQuestion)
    conds = []
    if topic:
        conds.append(QuizQuestion.topic == topic)
    if qtype:
        conds.append(QuizQuestion.qtype == qtype)
    if is_active is not None:
        conds.append(QuizQuestion.is_active.is_(is_active))
    if keyword:
        conds.append(QuizQuestion.stem.ilike(f"%{keyword}%"))
    if conds:
        stmt = stmt.where(and_(*conds))

    total = (
        await db.execute(
            select(func.count()).select_from(stmt.subquery())
        )
    ).scalar_one()

    rows = (
        await db.execute(
            stmt.order_by(QuizQuestion.id.desc())
            .offset((page - 1) * size)
            .limit(size)
        )
    ).scalars().all()
    return list(rows), int(total)


# ======================================================================
# D. 学生抽题 + 提交
# ======================================================================
async def draw_quiz(
    db: AsyncSession,
    *,
    topic: str | None,
    qtype: str | None,
    difficulty: str | None,
    limit: int,
) -> tuple[str, list[QuizQuestion]]:
    """随机抽取 limit 道有效题。返回 batch_id 与题目列表（不含答案）。"""
    stmt = select(QuizQuestion).where(QuizQuestion.is_active.is_(True))
    if topic:
        stmt = stmt.where(QuizQuestion.topic == topic)
    if qtype:
        stmt = stmt.where(QuizQuestion.qtype == qtype)
    if difficulty:
        stmt = stmt.where(QuizQuestion.difficulty == difficulty)

    rows = list((await db.execute(stmt)).scalars().all())
    if not rows:
        raise NotFoundError("无匹配题目，请联系管理员补充题库")
    random.shuffle(rows)
    picked = rows[: max(1, min(limit, len(rows)))]
    batch_id = uuid.uuid4().hex
    return batch_id, picked


async def submit_quiz(
    db: AsyncSession,
    *,
    student_id: int,
    batch_id: str,
    answers: list[dict],
    operator_id: int,
) -> dict:
    """整卷提交：逐题判分 + 落 quiz_records + 返回汇总。

    answers = [{"question_id": int, "answer": str}, ...]
    返回 {"batch_id", "total", "correct", "score", "items":[...]}。
    """
    if not answers:
        raise BizError("至少需要提交一道题")

    q_ids = [a["question_id"] for a in answers]
    rows = (
        await db.execute(select(QuizQuestion).where(QuizQuestion.id.in_(q_ids)))
    ).scalars().all()
    q_map = {q.id: q for q in rows}
    missing = set(q_ids) - set(q_map.keys())
    if missing:
        raise NotFoundError(f"题目不存在：{sorted(missing)}")

    now = datetime.now(UTC)
    per_score = 100 // len(answers) if len(answers) else 0
    items: list[dict] = []
    correct_count = 0

    for a in answers:
        q = q_map[a["question_id"]]
        answer_text = str(a.get("answer", ""))
        is_correct = grade_answer(q, answer_text)
        if is_correct:
            correct_count += 1
        record = QuizRecord(
            student_id=student_id,
            question_id=q.id,
            batch_id=batch_id,
            answer=answer_text,
            is_correct=is_correct,
            score=per_score if is_correct else 0,
            submitted_at=now,
        )
        db.add(record)
        items.append(
            {
                "question_id": q.id,
                "is_correct": is_correct,
                "correct_key": q.correct_key,
                "explanation": q.explanation,
            }
        )

    await db.flush()

    total = len(answers)
    score = correct_count * per_score
    await log_action(
        db,
        event_type="QUIZ",
        entity_code="QUIZ_BATCH",
        action="SUBMIT",
        actor_user_id=operator_id,
        detail={
            "batch_id": batch_id,
            "total": total,
            "correct": correct_count,
            "score": score,
        },
    )
    await db.commit()
    return {
        "batch_id": batch_id,
        "total": total,
        "correct": correct_count,
        "score": score,
        "items": items,
    }


async def list_my_records(
    db: AsyncSession,
    *,
    student_id: int,
    batch_id: str | None = None,
    limit: int = 100,
) -> list[QuizRecord]:
    stmt = select(QuizRecord).where(QuizRecord.student_id == student_id)
    if batch_id:
        stmt = stmt.where(QuizRecord.batch_id == batch_id)
    stmt = stmt.order_by(QuizRecord.submitted_at.desc()).limit(limit)
    rows = (await db.execute(stmt)).scalars().all()
    return list(rows)
