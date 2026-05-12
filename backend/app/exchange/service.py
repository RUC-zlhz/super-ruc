"""exchange 业务服务层 — FR-009 导入导出 / FR-015 培养方案维护。

关键原则（C-06）：
- 导入分两步：
  1) validate_and_store_batch: 解析 Excel，按行校验，入库 import_batches / import_batch_rows。
     - 不触碰正式表（students / student_course_records / curriculum_* ）。
  2) commit_batch: 如果没有 FATAL 行，则在单个事务内批量写入正式表；整批原子；失败回滚。

- 所有导入类型共用批次元信息 + 行级校验日志；具体解析逻辑按 import_type 分派。
"""
from __future__ import annotations

import io
import json
import logging
import uuid
from datetime import UTC, date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit.service import log_action
from app.core.exceptions import BizError, NotFoundError
from app.exchange import repository as repo
from app.exchange.models import (
    BATCH_STATUS_FAILED,
    BATCH_STATUS_VALIDATED,
    IMPORT_TYPE_COURSE_EQUIV,
    IMPORT_TYPE_COURSE_OFFERING,
    IMPORT_TYPE_CURRICULUM_MODULE,
    IMPORT_TYPE_HONOR,
    IMPORT_TYPE_STUDENT,
    IMPORT_TYPE_TRANSCRIPT,
    IMPORT_TYPE_TRANSCRIPT_PDF_REVIEW,
    ROW_RESULT_FAILED,
    ROW_SEVERITY_FATAL,
    ROW_SEVERITY_INFO,
    ROW_SEVERITY_WARN,
    ImportBatch,
)
from app.exchange.validators import get_validator

logger = logging.getLogger(__name__)


# ============================================================
# 工具
# ============================================================
def _gen_batch_no(import_type: str) -> str:
    return (
        f"IM-{import_type[:4].upper()}-"
        f"{datetime.now(UTC).strftime('%y%m%d%H%M%S')}-"
        f"{uuid.uuid4().hex[:6].upper()}"
    )


def _norm(v: Any) -> Any:
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _row_to_dict(header: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    return {header[i]: _norm(row[i]) for i in range(min(len(header), len(row)))}


# ============================================================
# 主入口：解析并落校验批次
# ============================================================
async def validate_excel_batch(
    db: AsyncSession,
    *,
    import_type: str,
    filename: str,
    file_bytes: bytes,
    object_bucket: str | None,
    object_key: str | None,
    operator_id: int,
    operator_role: str | None,
) -> ImportBatch:
    """读取 Excel → 逐行校验 → 创建 import_batches + import_batch_rows（VALIDATED/FAILED）。

    不写正式业务表。commit_batch 调用后才写正式表。
    """
    batch = await repo.create_batch(
        db,
        batch_no=_gen_batch_no(import_type),
        import_type=import_type,
        filename=filename,
        object_bucket=object_bucket,
        object_key=object_key,
        file_size=len(file_bytes),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        operator_id=operator_id,
        operator_role=operator_role,
    )

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        await repo.finalize_batch(
            db, batch,
            status=BATCH_STATUS_FAILED,
            total_rows=0, ok_rows=0, warn_rows=0, fatal_rows=1,
            summary={"error": f"Excel 读取失败: {e}"[:500]},
        )
        await log_action(
            db, event_type="IMPORT", entity_code="IMPORT_BATCH",
            action="VALIDATE", entity_id=batch.id,
            actor_user_id=operator_id, actor_role=operator_role,
            result_code="FAILED", message=str(e)[:256],
        )
        await db.commit()
        await db.refresh(batch)
        return batch

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        header_row = tuple()
    header = [str(h).strip() if h else "" for h in header_row]

    total = ok = warn = fatal = 0
    validator = get_validator(import_type)
    if validator is None:
        await repo.add_batch_row(
            db, batch_id=batch.id, row_no=0,
            severity=ROW_SEVERITY_FATAL, result=ROW_RESULT_FAILED,
            message=f"未知的 import_type: {import_type}",
        )
        fatal = 1
        total = 0
    else:
        for idx, raw in enumerate(rows_iter, start=2):
            total += 1
            if raw is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
                continue
            record = _row_to_dict(header, raw)
            sev, result, field, msg = validator(record)
            await repo.add_batch_row(
                db, batch_id=batch.id, row_no=idx,
                severity=sev, result=result,
                field_name=field, message=msg,
                raw_data=record,
            )
            if sev == ROW_SEVERITY_FATAL:
                fatal += 1
            elif sev == ROW_SEVERITY_WARN:
                warn += 1
                ok += 1
            else:
                ok += 1

    status = BATCH_STATUS_VALIDATED if fatal == 0 else BATCH_STATUS_FAILED
    await repo.finalize_batch(
        db, batch,
        status=status,
        total_rows=total, ok_rows=ok, warn_rows=warn, fatal_rows=fatal,
        summary={"header": header},
    )
    await log_action(
        db, event_type="IMPORT", entity_code="IMPORT_BATCH",
        action="VALIDATE", entity_id=batch.id,
        actor_user_id=operator_id, actor_role=operator_role,
        detail={"type": import_type, "total": total, "ok": ok,
                "warn": warn, "fatal": fatal},
    )
    await db.commit()
    await db.refresh(batch)
    return batch


# ============================================================
# Commit：把 VALIDATED 批次写入正式表（整批原子）
# ============================================================
async def commit_batch(
    db: AsyncSession,
    batch_id: int,
    *,
    note: str | None,
    operator_id: int,
    operator_role: str | None,
) -> ImportBatch:
    batch = await repo.get_batch(db, batch_id)
    if batch is None:
        raise NotFoundError("批次不存在")
    if batch.status != BATCH_STATUS_VALIDATED:
        raise BizError(f"批次状态 {batch.status} 不可提交，仅 VALIDATED 可提交", code=40040)
    if batch.fatal_rows > 0:
        raise BizError("批次存在致命错误，无法提交", code=40041)
    if batch.import_type == IMPORT_TYPE_TRANSCRIPT_PDF_REVIEW:
        raise BizError(
            "学生成绩单 PDF 上传记录仅供人工核验，不能通过导入提交写入正式成绩",
            code=40047,
        )

    applier = _APPLIERS.get(batch.import_type)
    if applier is None:
        raise BizError(f"未知的 import_type: {batch.import_type}", code=40042)

    try:
        await applier(db, batch)
    except Exception as e:  # noqa: BLE001
        # C-06 原子：任何异常回滚
        await db.rollback()
        fresh = await repo.get_batch(db, batch_id)
        if fresh is not None:
            await repo.mark_batch_rolled_back(db, fresh)
            await log_action(
                db, event_type="IMPORT", entity_code="IMPORT_BATCH",
                action="COMMIT", entity_id=fresh.id,
                actor_user_id=operator_id, actor_role=operator_role,
                result_code="FAILED", message=str(e)[:256],
            )
            await db.commit()
        raise BizError(f"导入提交失败：{e}", code=40043) from e

    await repo.mark_batch_committed(db, batch)
    if note is not None:
        batch.note = note
    await log_action(
        db, event_type="IMPORT", entity_code="IMPORT_BATCH",
        action="COMMIT", entity_id=batch.id,
        actor_user_id=operator_id, actor_role=operator_role,
        detail={"type": batch.import_type, "rows": batch.ok_rows},
    )
    await db.commit()
    await db.refresh(batch)
    return batch


# ============================================================
# 应用器：把已校验行写入正式表
# ============================================================
async def _apply_student(db: AsyncSession, batch: ImportBatch) -> None:
    rows = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_INFO, limit=100000)
    rows_warn = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_WARN, limit=100000)
    for r in rows + rows_warn:
        d = r.raw_data or {}
        payload = {
            "student_no": d.get("student_no"),
            "full_name": d.get("full_name"),
            "gender": d.get("gender"),
            "birth_date": _parse_date(d.get("birth_date")),
            "grade_code": d.get("grade_code"),
            "major_code": d.get("major_code"),
            "class_code": d.get("class_code"),
            "political_status": d.get("political_status"),
            "enrollment_year": _parse_int(d.get("enrollment_year")),
            "expected_graduation_year": _parse_int(d.get("expected_graduation_year")),
            "email": d.get("email"),
            "status": d.get("status") or "IN_SCHOOL",
        }
        await repo.upsert_student(db, {k: v for k, v in payload.items() if v is not None})


async def _apply_transcript(db: AsyncSession, batch: ImportBatch) -> None:
    rows = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_INFO, limit=100000)
    rows_warn = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_WARN, limit=100000)
    for r in rows + rows_warn:
        d = r.raw_data or {}
        student = await repo.get_student_by_no(db, str(d.get("student_no")))
        if student is None:
            raise BizError(f"学号 {d.get('student_no')} 不存在，需先导入学生主档", code=40044)
        payload = {
            "student_id": student.id,
            "term_code": d.get("term_code"),
            "course_code": d.get("course_code"),
            "course_name": d.get("course_name") or d.get("course_code"),
            "credits": float(d.get("credits") or 0),
            "course_type": d.get("course_type"),
            "score": _parse_float(d.get("score")),
            "grade_letter": d.get("grade_letter"),
            "pass_flag": _parse_bool(d.get("pass_flag"), default=False),
            "imported_batch_id": batch.id,
            "note": d.get("note"),
        }
        await repo.upsert_student_course_record(db, payload)


async def _apply_curriculum_module(db: AsyncSession, batch: ImportBatch) -> None:
    """按 (grade_code, major_code, plan_name) 分组，对每组清空 modules 后重建。"""
    rows = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_INFO, limit=100000)
    rows_warn = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_WARN, limit=100000)

    groups: dict[tuple[str, str, str | None], list[dict[str, Any]]] = {}
    plan_meta: dict[tuple[str, str, str | None], dict[str, Any]] = {}
    for r in rows + rows_warn:
        d = r.raw_data or {}
        key = (
            str(d.get("grade_code")),
            str(d.get("major_code")),
            d.get("version_label"),
        )
        groups.setdefault(key, []).append(d)
        plan_meta.setdefault(key, {
            "grade_code": d.get("grade_code"),
            "major_code": d.get("major_code"),
            "version_label": d.get("version_label"),
            "plan_name": d.get("plan_name") or f"{d.get('grade_code')}-{d.get('major_code')} 培养方案",
            "total_credits_required": _parse_float(d.get("total_credits_required")),
            "is_active": True,
        })

    for key, members in groups.items():
        plan = await repo.create_or_update_plan(db, plan_meta[key])
        modules = [
            {
                "module_code": m.get("module_code"),
                "module_name": m.get("module_name"),
                "module_type": m.get("module_type") or "REQUIRED",
                "credits_required": float(m.get("credits_required") or 0),
                "courses": _parse_courses(m.get("courses")),
                "note": m.get("note"),
                "sort_order": _parse_int(m.get("sort_order")) or 0,
            }
            for m in members
        ]
        await repo.set_plan_modules(db, plan.id, modules)


async def _apply_course_equiv(db: AsyncSession, batch: ImportBatch) -> None:
    rows = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_INFO, limit=100000)
    rows_warn = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_WARN, limit=100000)
    for r in rows + rows_warn:
        d = r.raw_data or {}
        payload = {
            "grade_code": d.get("grade_code"),
            "major_code": d.get("major_code"),
            "source_course_code": d.get("source_course_code"),
            "source_course_name": d.get("source_course_name"),
            "target_course_code": d.get("target_course_code"),
            "target_course_name": d.get("target_course_name"),
            "ratio": float(d.get("ratio") or 1.0),
            "note": d.get("note"),
            "is_active": _parse_bool(d.get("is_active"), default=True),
        }
        await repo.upsert_equivalence(db, payload)


async def _apply_course_offering(db: AsyncSession, batch: ImportBatch) -> None:
    rows = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_INFO, limit=100000)
    rows_warn = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_WARN, limit=100000)
    for r in rows + rows_warn:
        d = r.raw_data or {}
        payload = {
            "term_code": d.get("term_code"),
            "course_code": d.get("course_code"),
            "course_name": d.get("course_name"),
            "credits": float(d.get("credits") or 0),
            "course_type": d.get("course_type"),
            "major_codes": d.get("major_codes"),
            "grade_codes": d.get("grade_codes"),
            "teacher": d.get("teacher"),
            "capacity": _parse_int(d.get("capacity")),
            "note": d.get("note"),
            "is_active": _parse_bool(d.get("is_active"), default=True),
        }
        await repo.upsert_course_offering(db, payload)


async def _apply_honor(db: AsyncSession, batch: ImportBatch) -> None:
    from app.honor import repository as honor_repo
    from app.honor.models import HONOR_STATUS_ACTIVE

    rows = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_INFO, limit=100000)
    rows_warn = await repo.list_batch_rows(db, batch.id, severity=ROW_SEVERITY_WARN, limit=100000)
    grouped: dict[
        tuple[Any, ...], dict[str, Any]
    ] = {}
    for row in rows + rows_warn:
        data = row.raw_data or {}
        key = (
            data.get("category_code"),
            data.get("title"),
            data.get("level"),
            data.get("awarded_by"),
            data.get("document_no"),
            data.get("announced_at"),
            data.get("effective_from"),
            data.get("effective_to"),
            _parse_bool(data.get("is_collective"), default=False),
            data.get("summary"),
            data.get("story_md"),
            data.get("acceptance_speech"),
            data.get("cover_image_url"),
            _parse_bool(data.get("consent_flag"), default=True),
        )
        group = grouped.get(key)
        if group is None:
            group = {
                "payload": {
                    "category_code": data.get("category_code"),
                    "title": data.get("title"),
                    "level": data.get("level"),
                    "awarded_by": data.get("awarded_by"),
                    "document_no": data.get("document_no"),
                    "announced_at": _parse_date(data.get("announced_at")),
                    "effective_from": _parse_date(data.get("effective_from")),
                    "effective_to": _parse_date(data.get("effective_to")),
                    "is_collective": _parse_bool(data.get("is_collective"), default=False),
                    "summary": data.get("summary"),
                    "story_md": data.get("story_md"),
                    "acceptance_speech": data.get("acceptance_speech"),
                    "cover_image_url": data.get("cover_image_url"),
                    "consent_flag": _parse_bool(data.get("consent_flag"), default=True),
                    "status": HONOR_STATUS_ACTIVE,
                    "created_by": batch.operator_id,
                    "updated_by": batch.operator_id,
                },
                "recipients": [],
            }
            grouped[key] = group
        group["recipients"].append(data)

    categories = await honor_repo.get_categories_by_codes(
        db,
        {
            str(group["payload"]["category_code"])
            for group in grouped.values()
            if group["payload"].get("category_code")
        },
    )
    missing_codes = sorted(
        {
            str(group["payload"]["category_code"])
            for group in grouped.values()
            if group["payload"].get("category_code")
        }
        - set(categories.keys())
    )
    if missing_codes:
        raise BizError(f"荣誉类别不存在：{', '.join(missing_codes)}", code=40045)

    for group in grouped.values():
        record = await honor_repo.create_record(db, group["payload"])
        recipients: list[dict[str, Any]] = []
        for raw in group["recipients"]:
            student = None
            student_no = str(raw.get("student_no") or raw.get("student_no_snapshot") or "").strip()
            if student_no:
                student = await repo.get_student_by_no(db, student_no)
                if student is None:
                    raise BizError(f"学号 {student_no} 未在学生主档中，无法核验荣誉记录", code=40046)
            recipients.append(
                {
                    "student_id": student.id if student is not None else None,
                    "student_no_snapshot": student.student_no if student is not None else (student_no or None),
                    "display_name": (
                        raw.get("display_name")
                        or (student.full_name if student is not None else None)
                        or student_no
                    ),
                    "major_snapshot": raw.get("major_snapshot") or (student.major_code if student is not None else None),
                    "grade_snapshot": raw.get("grade_snapshot") or (student.grade_code if student is not None else None),
                    "class_snapshot": raw.get("class_snapshot") or (student.class_code if student is not None else None),
                    "role_in_collective": raw.get("role_in_collective"),
                }
            )
        await honor_repo.set_recipients(db, record.id, recipients)


_APPLIERS = {
    IMPORT_TYPE_STUDENT: _apply_student,
    IMPORT_TYPE_TRANSCRIPT: _apply_transcript,
    IMPORT_TYPE_CURRICULUM_MODULE: _apply_curriculum_module,
    IMPORT_TYPE_COURSE_EQUIV: _apply_course_equiv,
    IMPORT_TYPE_COURSE_OFFERING: _apply_course_offering,
    IMPORT_TYPE_HONOR: _apply_honor,
}


# ============================================================
# 类型转换辅助
# ============================================================
def _parse_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _parse_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_bool(v: Any, *, default: bool = False) -> bool:
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "是", "通过", "合格")


def _parse_courses(v: Any) -> list[dict[str, Any]] | None:
    if v in (None, ""):
        return None
    if isinstance(v, list):
        return [item for item in v if isinstance(item, dict)] or None
    if isinstance(v, dict):
        return [v]
    if isinstance(v, str):
        try:
            parsed = json.loads(v)
        except json.JSONDecodeError:
            return None
        return _parse_courses(parsed)
    return None


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    try:
        return datetime.strptime(str(v), "%Y-%m-%d").date()
    except ValueError:
        return None


# ============================================================
# v1.5：批次错误报告 Excel
# ============================================================
ERROR_REPORT_COLUMN = "错误原因"


async def build_error_report(
    db: AsyncSession, batch_id: int,
) -> tuple[bytes, str]:
    """为失败或带警告的批次生成带"错误原因"列的 Excel（FR-009 / v1.5）。

    - 列顺序：原 Excel 首行（从 summary.header 恢复）+ 末尾追加 `错误原因`。
    - 行顺序：按 row_no 升序；每行 raw_data 按 header 展开。
    - 只输出 FATAL/WARN 行；INFO/OK 行跳过（如需全量导出用 imports/{id}）。
    - 文件名形如 `import-errors-<batch_no>.xlsx`。
    """
    batch = await repo.get_batch(db, batch_id)
    if batch is None:
        raise NotFoundError("批次不存在")

    header: list[str] = []
    if isinstance(batch.summary, dict):
        header = [str(h) for h in batch.summary.get("header") or []]
    # 若 summary 里没存 header，退化为 raw_data 中出现过的键
    if not header:
        sample_rows = await repo.list_batch_rows(db, batch_id, limit=5)
        seen: list[str] = []
        for r in sample_rows:
            for k in (r.raw_data or {}).keys():
                if k not in seen:
                    seen.append(k)
        header = seen

    err_rows = []
    for sev in (ROW_SEVERITY_FATAL, ROW_SEVERITY_WARN):
        err_rows.extend(await repo.list_batch_rows(db, batch_id, severity=sev, limit=100000))
    err_rows.sort(key=lambda r: r.row_no)

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    ws.append(header + [ERROR_REPORT_COLUMN])
    for r in err_rows:
        data = r.raw_data or {}
        line = [_cell_value(data.get(h)) for h in header]
        reason_parts: list[str] = []
        if r.field_name:
            reason_parts.append(f"[{r.field_name}]")
        if r.message:
            reason_parts.append(r.message)
        reason_parts.append(f"(row {r.row_no}, {r.severity})")
        line.append(" ".join(reason_parts))
        ws.append(line)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"import-errors-{batch.batch_no}.xlsx"


# ============================================================
# 导出：构造 Excel bytes（供 router 以 StreamingResponse 返回）
# ============================================================
def export_workbook(headers: list[str], rows: list[list[Any]], sheet_name: str = "export") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "export"
    ws.append(headers)
    for r in rows:
        ws.append([_cell_value(c) for c in r])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_value(v: Any) -> Any:
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, (list, dict)):
        return str(v)
    return v


# ============================================================
# 培养方案 / 课程等价 / 开课 的 CRUD 服务层封装
# ============================================================
async def upsert_plan_with_modules(
    db: AsyncSession,
    payload: dict[str, Any],
    *,
    operator_id: int,
    operator_role: str | None,
) -> int:
    modules = payload.pop("modules", []) or []
    plan = await repo.create_or_update_plan(db, payload)
    await repo.set_plan_modules(
        db, plan.id,
        [m.model_dump() if hasattr(m, "model_dump") else m for m in modules],
    )
    await log_action(
        db, event_type="CURRICULUM", entity_code="CURRICULUM_PLAN",
        action="UPSERT", entity_id=plan.id,
        actor_user_id=operator_id, actor_role=operator_role,
        detail={"grade_code": plan.grade_code, "major_code": plan.major_code,
                "module_count": len(modules)},
    )
    await db.commit()
    await db.refresh(plan)
    return plan.id


async def delete_plan(
    db: AsyncSession, plan_id: int, *,
    operator_id: int, operator_role: str | None,
) -> None:
    plan = await repo.get_plan(db, plan_id)
    if plan is None:
        raise NotFoundError("培养方案不存在")
    await repo.delete_plan(db, plan)
    await log_action(
        db, event_type="CURRICULUM", entity_code="CURRICULUM_PLAN",
        action="DELETE", entity_id=plan_id,
        actor_user_id=operator_id, actor_role=operator_role,
    )
    await db.commit()
