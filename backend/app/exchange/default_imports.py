"""S12 默认学生与培养方案导入服务。

默认导入只消费仓库内已登记的数据源，不从学号或毕业年份推断专业/班级。
培养方案导入只维护 version_label=2024-default 的演示版本，不覆盖教师后续维护版本。
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import UTC, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.audit.service import log_action
from app.core.exceptions import BizError
from app.exchange import repository as repo
from app.exchange.models import CurriculumPlan
from app.exchange.schemas import DefaultImportResult

_REPO_ROOT = Path(__file__).resolve().parents[3]
_DEFAULT_STUDENTS_PATH = _REPO_ROOT / "docs" / "source" / "students" / "students.xlsx"
_DEFAULT_CURRICULUM_PATH = (
    _REPO_ROOT / "docs" / "source" / "training program" / "2024_information.md"
)
_DEFAULT_GRADE_CODE = "2024"
_DEFAULT_VERSION_LABEL = "2024-default"
_SOURCE_COVERAGE_MAJOR_CODE = "源文件全量课程池"
_TARGET_MAJORS = [
    "计算机科学与技术专业",
    "信息管理与信息系统专业",
    "软件工程专业",
    "信息安全专业",
    "数据科学与大数据技术专业",
    "数据科学与大数据技术（理学）专业",
]
_PLAN_TOTAL_CREDITS = {
    "计算机科学与技术": 155.0,
    "信息管理与信息系统": 153.0,
    "软件工程": 156.0,
    "信息安全": 155.0,
    "数据科学与大数据技术": 155.0,
    "数据科学与大数据技术（理学）": 153.0,
}
_PERSONALIZED_ELECTIVE_CREDITS = {
    "计算机科学与技术": 18.0,
    "信息管理与信息系统": 22.0,
    "软件工程": 18.0,
    "信息安全": 18.0,
    "数据科学与大数据技术": 19.0,
    "数据科学与大数据技术（理学）": 16.0,
}
_PROFESSIONAL_CORE_CREDITS = {
    "计算机科学与技术": 46.0,
    "信息管理与信息系统": 40.0,
    "软件工程": 47.0,
    "信息安全": 46.0,
    "数据科学与大数据技术": 45.0,
    "数据科学与大数据技术（理学）": 46.0,
}
_COMMON_CREDIT_OVERRIDES = (
    ("思想政治理论课", 21.0),
    ("公共外语", 6.0),
    ("公共体育", 4.0),
    ("新生研讨课", 2.0),
    ("心理健康教育", 2.0),
    ("部类核心课", 26.0),
    ("研究训练", 2.0),
    ("专业实习", 4.0),
    ("毕业论文", 4.0),
    ("劳动教育", 1.0),
    ("军事课", 4.0),
    ("职业生涯规划", 1.0),
    ("志愿服务", 2.0),
)
_REQUIREMENT_ONLY_MODULES = (
    ("GENERAL-ELECTIVE", "通识课程群（通识核心课、一般通识课）", "GENERAL", 6.0),
    ("AESTHETIC", "美育课程", "GENERAL", 2.0),
    ("INTERNATIONAL-SUMMER", "国际暑期学校全英文课", "GENERAL", 2.0),
    ("PUBLIC-ELECTIVE", "公共选修课", "ELECTIVE", 2.0),
)
_DATA_SCIENCE_SCIENCE_FEATURE_CODES = {
    "BSTAMS0030S",
    "BSTAMS0022",
    "BBSMMSB005",
    "BSTAMS0010S",
    "BBSMMS0007",
    "BPTMMS0004",
    "BORCMS0004S",
    "BSTAMS0026S",
}
_DATA_SCIENCE_ENGINEERING_FEATURE_CODES = {
    "BBSEMS0006S",
    "BCSAMSS0005S",
    "BCSAMSB001S",
    "BCSTMSA004",
    "BCSTMS0002S",
    "BCSTMSB007S",
    "BBSEMS0028S",
    "BISYMS0005S",
    "BCSTMS0008S",
}
_COURSE_CODE_RE = re.compile(r"^(?=.*[A-Za-z])(?=.*\d)[A-Za-z][A-Za-z0-9_-]{4,}$")
_NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")


@dataclass(slots=True)
class _Cell:
    text: str
    rowspan: int = 1
    colspan: int = 1


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[_Cell]] = []
        self._current_row: list[_Cell] | None = None
        self._current_cell: dict[str, Any] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        lowered = tag.lower()
        if lowered == "tr":
            self._current_row = []
        elif lowered in {"td", "th"} and self._current_row is not None:
            attr_map = dict(attrs)
            self._current_cell = {
                "parts": [],
                "rowspan": _safe_span(attr_map.get("rowspan")),
                "colspan": _safe_span(attr_map.get("colspan")),
            }

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell["parts"].append(data)

    def handle_endtag(self, tag: str) -> None:
        lowered = tag.lower()
        if lowered in {"td", "th"} and self._current_cell is not None:
            text = _clean_text("".join(self._current_cell["parts"]))
            self._current_row.append(
                _Cell(
                    text=text,
                    rowspan=self._current_cell["rowspan"],
                    colspan=self._current_cell["colspan"],
                )
            )
            self._current_cell = None
        elif lowered == "tr" and self._current_row is not None:
            self.rows.append(self._current_row)
            self._current_row = None


def _safe_span(value: str | None) -> int:
    try:
        return max(1, int(value or "1"))
    except ValueError:
        return 1


def _clean_text(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()


def _normalize_student_no(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = _clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _parse_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _parse_float(value: Any) -> float | None:
    match = _NUMBER_RE.search(_clean_text(value))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _looks_like_course_code(value: str) -> bool:
    code = _clean_text(value).replace(" ", "")
    if code in {"", "/", "\\", "—", "-", "课程编码", "课程代码"}:
        return False
    return bool(_COURSE_CODE_RE.match(code))


def _matrix_from_html_table(html: str) -> list[list[str]]:
    parser = _TableParser()
    parser.feed(html)
    pending: dict[int, tuple[str, int]] = {}
    matrix: list[list[str]] = []

    def fill_pending(row: list[str], col: int) -> int:
        while col in pending:
            text, rows_left = pending[col]
            while len(row) <= col:
                row.append("")
            row[col] = text
            if rows_left <= 1:
                pending.pop(col, None)
            else:
                pending[col] = (text, rows_left - 1)
            col += 1
        return col

    for raw_row in parser.rows:
        row: list[str] = []
        col = 0
        for cell in raw_row:
            col = fill_pending(row, col)
            for offset in range(cell.colspan):
                target_col = col + offset
                while len(row) <= target_col:
                    row.append("")
                row[target_col] = cell.text
                if cell.rowspan > 1:
                    pending[target_col] = (cell.text, cell.rowspan - 1)
            col += cell.colspan
        fill_pending(row, col)
        matrix.append(row)
    return matrix


def _find_column(row: list[str], aliases: tuple[str, ...]) -> int | None:
    for index, value in enumerate(row):
        text = _clean_text(value)
        if any(alias in text for alias in aliases):
            return index
    return None


def _find_course_header(matrix: list[list[str]]) -> tuple[int, int, int, int, int] | None:
    for row_index, row in enumerate(matrix[:8]):
        name_idx = _find_column(row, ("课程名称", "课程名"))
        code_idx = _find_column(row, ("课程编码", "课程代码", "编码"))
        credits_idx = _find_column(row, ("学分",))
        term_idx = _find_column(row, ("开课学期", "学期"))
        if name_idx is not None and code_idx is not None and credits_idx is not None:
            return row_index, name_idx, code_idx, credits_idx, term_idx if term_idx is not None else -1
    return None


def _course_header_for_row(row: list[str]) -> tuple[int, int, int, int, int | None] | None:
    name_idx = _find_column(row, ("课程名称", "课程名"))
    code_idx = _find_column(row, ("课程编码", "课程代码", "编码"))
    credits_idx = _find_column(row, ("学分",))
    term_idx = _find_column(row, ("开课学期", "学期"))
    major_idx = _find_column(row, ("专业名称",))
    if name_idx is not None and code_idx is not None and credits_idx is not None:
        return name_idx, code_idx, credits_idx, term_idx if term_idx is not None else -1, major_idx
    return None


def _last_context_title(section: str, table_start: int) -> str:
    before = section[:table_start].splitlines()
    for line in reversed(before):
        text = _clean_text(line).lstrip("#").strip()
        if text and not text.startswith("<"):
            return text[:96]
    return "课程模块"


def _module_type(title: str) -> str:
    if "通识" in title:
        return "GENERAL"
    if any(token in title for token in ("实践", "实习", "论文", "训练")):
        return "PRACTICE"
    if "选修" in title:
        return "ELECTIVE"
    return "REQUIRED"


def _module_code(scope: str, title: str, seq: int) -> str:
    digest = hashlib.sha256(
        f"{scope}|{title}|{seq}".encode(),
        usedforsecurity=False,
    ).hexdigest()[:8]
    normalized_scope = re.sub(r"[^A-Za-z0-9]+", "-", scope.upper()).strip("-") or "MODULE"
    return f"{normalized_scope[:24]}-{seq:03d}-{digest}"


def _extract_course_modules(
    section: str,
    *,
    scope: str,
    major_name: str | None = None,
    include_shared_without_major: bool = False,
    include_all_major_rows: bool = False,
) -> list[dict[str, Any]]:
    modules: list[dict[str, Any]] = []
    for seq, match in enumerate(re.finditer(r"<table>.*?</table>", section, flags=re.S), start=1):
        title = _last_context_title(section, match.start())
        matrix = _matrix_from_html_table(match.group(0))
        normalized_major_name = _major_alias_key(major_name) if major_name else None
        courses: list[dict[str, Any]] = []
        seen_codes: set[str] = set()
        current_header: tuple[int, int, int, int, int | None] | None = None
        for row in matrix:
            next_header = _course_header_for_row(row)
            if next_header is not None:
                current_header = next_header
                continue
            if current_header is None:
                continue
            name_idx, code_idx, credits_idx, term_idx, major_idx = current_header
            required_indices = [name_idx, code_idx, credits_idx]
            if term_idx >= 0:
                required_indices.append(term_idx)
            if max(required_indices) >= len(row):
                continue
            if major_idx is not None:
                if major_name is None and not include_all_major_rows:
                    continue
                if major_idx >= len(row) and not include_all_major_rows:
                    continue
                row_major = _major_alias_key(row[major_idx]) if major_idx < len(row) else ""
                if major_name is not None and row_major != normalized_major_name:
                    continue
            elif major_name is not None and not include_shared_without_major:
                continue
            name = _clean_text(row[name_idx])
            code = _clean_text(row[code_idx]).replace(" ", "")
            if not _looks_like_course_code(code):
                continue
            if name in {"", "课程名称"} or code in seen_codes:
                continue
            seen_codes.add(code)
            credits = _parse_float(row[credits_idx])
            course: dict[str, Any] = {"code": code, "name": name}
            if credits is not None:
                course["credits"] = credits
            if term_idx >= 0 and term_idx < len(row):
                term = _clean_text(row[term_idx])
                if term and term not in {"开课学期", "/"}:
                    course["opening_term"] = term
            courses.append(course)
        if not courses:
            continue
        credits_required = round(
            sum(float(course.get("credits") or 0) for course in courses),
            2,
        )
        modules.append(
            {
                "module_code": _module_code(scope, title, seq),
                "module_name": title,
                "module_type": _module_type(title),
                "credits_required": credits_required,
                "courses": courses,
                "note": "默认导入：按培养方案表格课程清单生成，需教师复核后作为正式版本使用。",
                "sort_order": seq,
            }
        )
    return modules


def _extract_source_coverage_modules(text: str) -> list[dict[str, Any]]:
    modules = _extract_course_modules(
        text,
        scope="SOURCE",
        include_all_major_rows=True,
    )
    normalized = _merge_same_named_modules(modules)
    for sort_order, module in enumerate(normalized, start=1):
        module["module_code"] = _module_code(
            "SOURCE",
            str(module.get("module_name") or "课程模块"),
            sort_order,
        )
        module["module_type"] = str(module.get("module_type") or "OTHER")
        module["credits_required"] = round(
            sum(float(course.get("credits") or 0) for course in module.get("courses") or []),
            2,
        )
        module["note"] = "默认导入：源文件全量课程池，用于完整保存 2024_information.md 中可解析课程。"
        module["sort_order"] = sort_order
    return normalized


def _normalize_default_modules(
    modules: list[dict[str, Any]],
    *,
    major_code: str,
) -> list[dict[str, Any]]:
    modules = _filter_data_science_variant_modules(modules, major_code=major_code)
    modules = _merge_same_named_modules([*modules, *_requirement_only_modules(major_code)])
    normalized: list[dict[str, Any]] = []
    for sort_order, module in enumerate(modules, start=1):
        item = dict(module)
        title = str(item.get("module_name") or "")
        for keyword, credits in _COMMON_CREDIT_OVERRIDES:
            if keyword in title:
                item["credits_required"] = credits
                break
        if "个性化选修课" in title:
            item["credits_required"] = _PERSONALIZED_ELECTIVE_CREDITS.get(
                major_code,
                float(item.get("credits_required") or 0),
            )
            item["module_type"] = "ELECTIVE"
        if "专业核心课" in title:
            item["credits_required"] = _PROFESSIONAL_CORE_CREDITS.get(
                major_code,
                float(item.get("credits_required") or 0),
            )
        item["sort_order"] = sort_order
        normalized.append(item)
    return normalized


def _filter_data_science_variant_modules(
    modules: list[dict[str, Any]],
    *,
    major_code: str,
) -> list[dict[str, Any]]:
    if major_code not in {"数据科学与大数据技术", "数据科学与大数据技术（理学）"}:
        return modules
    feature_codes = (
        _DATA_SCIENCE_SCIENCE_FEATURE_CODES
        if major_code == "数据科学与大数据技术（理学）"
        else _DATA_SCIENCE_ENGINEERING_FEATURE_CODES
    )
    filtered: list[dict[str, Any]] = []
    for module in modules:
        item = dict(module)
        courses = list(item.get("courses") or [])
        if item.get("module_name") == "2. 专业核心课":
            variant_courses = [
                course
                for course in courses
                if str(course.get("code") or "") not in (
                    _DATA_SCIENCE_SCIENCE_FEATURE_CODES
                    | _DATA_SCIENCE_ENGINEERING_FEATURE_CODES
                )
                or str(course.get("code") or "") in feature_codes
            ]
            item["courses"] = variant_courses
            item["credits_required"] = round(
                sum(float(course.get("credits") or 0) for course in variant_courses),
                2,
            )
        filtered.append(item)
    return filtered


def _requirement_only_modules(major_code: str) -> list[dict[str, Any]]:
    return [
        {
            "module_code": f"REQ-{code}",
            "module_name": name,
            "module_type": module_type,
            "credits_required": credits,
            "courses": [],
            "note": (
                "默认导入：源培养方案仅给出最低学分要求，未提供固定课程编码；"
                "需教师维护课程白名单或以人工核验为准。"
            ),
            "sort_order": 900 + index,
        }
        for index, (code, name, module_type, credits) in enumerate(
            _REQUIREMENT_ONLY_MODULES,
            start=1,
        )
    ]


def _merge_same_named_modules(modules: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    by_name: dict[str, dict[str, Any]] = {}
    for module in modules:
        name = str(module.get("module_name") or "")
        if name not in by_name:
            item = dict(module)
            item["courses"] = list(module.get("courses") or [])
            by_name[name] = item
            merged.append(item)
            continue
        item = by_name[name]
        seen_codes = {
            str(course.get("code") or "")
            for course in item.get("courses") or []
            if isinstance(course, dict)
        }
        for course in module.get("courses") or []:
            if not isinstance(course, dict):
                continue
            code = str(course.get("code") or "")
            if not code or code in seen_codes:
                continue
            item.setdefault("courses", []).append(course)
            seen_codes.add(code)
        item["credits_required"] = round(
            sum(float(course.get("credits") or 0) for course in item.get("courses") or []),
            2,
        )
    return merged


def _normalize_major_name(major_heading: str) -> str:
    return _clean_text(major_heading).removesuffix("专业")


def _major_alias_key(value: str) -> str:
    text = _clean_text(value).removesuffix("专业")
    text = re.sub(r"\s*\^.*$", "", text)
    return text


def _split_major_sections(text: str) -> tuple[str, dict[str, str]]:
    positions: list[tuple[str, int]] = []
    for major in _TARGET_MAJORS:
        marker = f"# {major}"
        index = text.find(marker)
        if index >= 0:
            positions.append((major, index))
    positions.sort(key=lambda item: item[1])
    if not positions:
        raise BizError("默认培养方案中未找到目标专业标题", code=40090)

    common = text[: positions[0][1]]
    sections: dict[str, str] = {}
    for idx, (major, start) in enumerate(positions):
        end = positions[idx + 1][1] if idx + 1 < len(positions) else len(text)
        second_part = text.find("\n# 第二部分", start, end)
        if second_part >= 0:
            end = second_part
        sections[major] = text[start:end]
    return common, sections


async def _get_default_plan(
    db: AsyncSession, *, grade_code: str, major_code: str
) -> CurriculumPlan | None:
    stmt = select(CurriculumPlan).where(
        CurriculumPlan.grade_code == grade_code,
        CurriculumPlan.major_code == major_code,
        CurriculumPlan.version_label == _DEFAULT_VERSION_LABEL,
    )
    return (await db.execute(stmt)).scalar_one_or_none()


async def import_default_students(
    db: AsyncSession,
    *,
    operator_id: int | None,
    operator_role: str | None,
    source_path: Path = _DEFAULT_STUDENTS_PATH,
    commit: bool = True,
) -> DefaultImportResult:
    if not source_path.exists():
        raise BizError(f"默认学生花名册不存在：{source_path}", code=40091)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [_clean_text(value) for value in next(rows)]
    except StopIteration as exc:
        raise BizError("默认学生花名册为空", code=40092) from exc
    index = {name: pos for pos, name in enumerate(header)}
    required = {"学号", "姓名"}
    missing = sorted(required - set(index))
    if missing:
        raise BizError(f"默认学生花名册缺少列：{', '.join(missing)}", code=40093)

    created = updated = skipped = total = 0
    warnings: list[str] = []
    for raw_row in rows:
        if raw_row is None or all(value in (None, "") for value in raw_row):
            continue
        total += 1
        student_no = _normalize_student_no(raw_row[index["学号"]])
        full_name = _clean_text(raw_row[index["姓名"]])
        if not student_no or not full_name:
            skipped += 1
            warnings.append(f"第 {total + 1} 行缺少学号或姓名，已跳过。")
            continue
        payload = {
            "student_no": student_no,
            "full_name": full_name,
            "gender": _clean_text(raw_row[index["性别"]]) if "性别" in index else None,
            "expected_graduation_year": _parse_int(raw_row[index["毕业年份"]])
            if "毕业年份" in index
            else None,
        }
        _, was_created = await repo.upsert_student(
            db, {key: value for key, value in payload.items() if value is not None}
        )
        if was_created:
            created += 1
        else:
            updated += 1

    await log_action(
        db,
        event_type="IMPORT",
        entity_code="DEFAULT_STUDENT_IMPORT",
        action="DEFAULT_IMPORT",
        actor_user_id=operator_id,
        actor_role=operator_role,
        detail={
            "source": str(source_path),
            "total": total,
            "created": created,
            "updated": updated,
            "skipped": skipped,
        },
    )
    if commit:
        await db.commit()
    return DefaultImportResult(
        import_type="DEFAULT_STUDENTS",
        total_rows=total,
        created_count=created,
        updated_count=updated,
        skipped_count=skipped,
        warning_count=len(warnings),
        warnings=warnings[:20],
    )


async def import_default_curriculum(
    db: AsyncSession,
    *,
    operator_id: int | None,
    operator_role: str | None,
    source_path: Path = _DEFAULT_CURRICULUM_PATH,
    commit: bool = True,
) -> DefaultImportResult:
    if not source_path.exists():
        raise BizError(f"默认培养方案不存在：{source_path}", code=40094)
    text = source_path.read_text(encoding="utf-8")
    professional_marker = "# （二）专业模块"
    professional_index = text.find(professional_marker)
    if professional_index < 0:
        raise BizError("默认培养方案中未找到专业模块分隔符", code=40096)
    common_text = text[:professional_index]
    major_heading_index = text.find(f"# {_TARGET_MAJORS[0]}", professional_index)
    if major_heading_index < 0:
        major_heading_index = len(text)
    professional_text = text[professional_index:major_heading_index]
    common_modules = _extract_course_modules(common_text, scope="COMMON")
    if not common_modules:
        raise BizError("默认培养方案未解析出共享课程模块", code=40095)

    created = updated = skipped = total = 0
    warnings: list[str] = []
    for major_heading in _TARGET_MAJORS:
        major_code = _normalize_major_name(major_heading)
        course_table_major_name = (
            "数据科学与大数据技术"
            if major_code == "数据科学与大数据技术（理学）"
            else major_code
        )
        major_modules = _extract_course_modules(
            professional_text,
            scope=major_code,
            major_name=course_table_major_name,
            include_shared_without_major=True,
        )
        if not major_modules:
            skipped += 1
            warnings.append(f"未找到专业课程模块：{major_heading}")
            continue
        modules = _normalize_default_modules(
            [dict(module) for module in common_modules] + major_modules,
            major_code=major_code,
        )
        if not modules:
            skipped += 1
            warnings.append(f"专业 {major_heading} 未解析出课程模块")
            continue
        total += 1
        existing = await _get_default_plan(
            db,
            grade_code=_DEFAULT_GRADE_CODE,
            major_code=major_code,
        )
        plan = await repo.create_or_update_plan(
            db,
            {
                "grade_code": _DEFAULT_GRADE_CODE,
                "major_code": major_code,
                "plan_name": f"{_DEFAULT_GRADE_CODE}级{major_code}专业培养方案（默认）",
                "version_label": _DEFAULT_VERSION_LABEL,
                "total_credits_required": _PLAN_TOTAL_CREDITS.get(major_code),
                "effective_from": datetime(2024, 9, 1, tzinfo=UTC).date(),
                "is_active": True,
                "note": (
                    "S12 默认导入版本，仅作为初始演示数据；"
                    "教师后续维护的非默认版本不受本导入覆盖。"
                ),
            },
        )
        await repo.set_plan_modules(db, plan.id, modules)
        if existing is None:
            created += 1
        else:
            updated += 1

    source_coverage_modules = _extract_source_coverage_modules(text)
    if not source_coverage_modules:
        skipped += 1
        warnings.append("未解析出源文件全量课程池")
    else:
        total += 1
        existing = await _get_default_plan(
            db,
            grade_code=_DEFAULT_GRADE_CODE,
            major_code=_SOURCE_COVERAGE_MAJOR_CODE,
        )
        plan = await repo.create_or_update_plan(
            db,
            {
                "grade_code": _DEFAULT_GRADE_CODE,
                "major_code": _SOURCE_COVERAGE_MAJOR_CODE,
                "plan_name": f"{_DEFAULT_GRADE_CODE}级培养方案源文件全量课程池（默认）",
                "version_label": _DEFAULT_VERSION_LABEL,
                "total_credits_required": None,
                "effective_from": datetime(2024, 9, 1, tzinfo=UTC).date(),
                "is_active": False,
                "note": (
                    "S19 默认导入补充记录，用于完整保存 "
                    "docs/source/training program/2024_information.md 中可解析课程；"
                    "不参与学生学业缺口自动匹配。"
                ),
            },
        )
        await repo.set_plan_modules(db, plan.id, source_coverage_modules)
        if existing is None:
            created += 1
        else:
            updated += 1

    await log_action(
        db,
        event_type="IMPORT",
        entity_code="DEFAULT_CURRICULUM_IMPORT",
        action="DEFAULT_IMPORT",
        actor_user_id=operator_id,
        actor_role=operator_role,
        detail={
            "source": str(source_path),
            "grade_code": _DEFAULT_GRADE_CODE,
            "version_label": _DEFAULT_VERSION_LABEL,
            "plans": total,
            "created": created,
            "updated": updated,
            "skipped": skipped,
        },
    )
    if commit:
        await db.commit()
    return DefaultImportResult(
        import_type="DEFAULT_CURRICULUM",
        total_rows=total,
        created_count=created,
        updated_count=updated,
        skipped_count=skipped,
        warning_count=len(warnings),
        warnings=warnings[:20],
    )
