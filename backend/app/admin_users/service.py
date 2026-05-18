"""Backend-account bulk import service."""
from __future__ import annotations

import csv
import io
import re
import secrets
import string
import uuid
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.admin_users import repository as repo
from app.admin_users.models import (
    BATCH_STATUS_FAILED,
    BATCH_STATUS_VALIDATED,
    ROW_RESULT_CREATED,
    ROW_RESULT_EXISTING,
    ROW_RESULT_FAILED,
    ROW_RESULT_ROLE_GRANTED,
    ROW_RESULT_VALID,
    ROW_SEVERITY_FATAL,
    ROW_SEVERITY_INFO,
    ROW_SEVERITY_WARN,
    AdminUserImportBatch,
)
from app.admin_users.schemas import AdminUserCredentialOut
from app.audit.service import build_audit_detail, log_action
from app.auth import repository as auth_repo
from app.auth.models import Role, User
from app.auth.role_codes import normalize_role_code
from app.auth.scopes import (
    SCOPE_GLOBAL,
    VALID_SCOPE_TYPES,
    build_scope_code,
    normalize_scope_type,
)
from app.core.exceptions import BizError, NotFoundError, PermissionError
from app.core.security import hash_password

TEMPLATE_COLUMNS = [
    "work_no",
    "display_name",
    "email",
    "role_code",
    "scope_type",
    "scope_code",
    "is_active",
]
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

IMPORTER_ROLES = {
    "SUPER_ADMIN",
    "COLLEGE_LEADER",
    "COUNSELOR",
    "HEAD_TEACHER",
    "YOUTH_LEAGUE_TEACHER",
    "PARTY_BUILD_TEACHER",
}
L3_IMPORTER_ROLES = {
    "COUNSELOR",
    "HEAD_TEACHER",
    "YOUTH_LEAGUE_TEACHER",
    "PARTY_BUILD_TEACHER",
}
DISALLOWED_TARGET_ROLES = {"STUDENT", "GUEST"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
SYMBOL_ALPHABET = "!#%*?"


def ensure_import_permission(actor_roles: list[str]) -> None:
    if not set(actor_roles) & IMPORTER_ROLES:
        raise PermissionError("当前角色不可批量创建后台账号")


def _actor_role_label(actor_roles: list[str]) -> str | None:
    return ",".join(actor_roles) or None


def _actor_can_create_role(actor_roles: list[str], role: Role, scope_code: str | None) -> tuple[bool, str | None]:
    role_code = normalize_role_code(role.code) or role.code
    if role_code in DISALLOWED_TARGET_ROLES:
        return False, "后台账号导入不支持创建学生/访客账号"
    if "SUPER_ADMIN" in actor_roles:
        return True, None
    if "COLLEGE_LEADER" in actor_roles:
        if role.level in {3, 4} and role_code not in {"SUPER_ADMIN", "COLLEGE_LEADER"}:
            return True, None
        return False, "学院领导仅可创建 L3/L4 后台账号"
    if set(actor_roles) & L3_IMPORTER_ROLES:
        if role.level != 4:
            return False, "L3 老师仅可创建 L4 学生骨干后台账号"
        if not scope_code:
            return False, "L3 老师创建 L4 账号时必须填写 CLASS/MAJOR/GRADE 范围"
        return True, None
    return False, "当前角色不可批量创建后台账号"


def _gen_batch_no() -> str:
    return (
        f"AU-{datetime.now(UTC).strftime('%y%m%d%H%M%S')}-"
        f"{uuid.uuid4().hex[:8].upper()}"
    )


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def _parse_bool(raw: str | None) -> bool | None:
    value = (raw or "").strip().lower()
    if value == "":
        return True
    if value in {"1", "true", "yes", "y", "active", "enabled", "启用", "是"}:
        return True
    if value in {"0", "false", "no", "n", "inactive", "disabled", "停用", "否"}:
        return False
    return None


def _generate_initial_password() -> str:
    rng = secrets.SystemRandom()
    chars = [
        rng.choice(string.ascii_lowercase),
        rng.choice(string.ascii_uppercase),
        rng.choice(string.digits),
        rng.choice(SYMBOL_ALPHABET),
    ]
    alphabet = string.ascii_letters + string.digits + SYMBOL_ALPHABET
    chars.extend(rng.choice(alphabet) for _ in range(10))
    rng.shuffle(chars)
    return "".join(chars)


def _read_csv_rows(file_bytes: bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk")
    reader = csv.DictReader(io.StringIO(text))
    header = [str(h).strip() for h in (reader.fieldnames or [])]
    rows: list[tuple[int, dict[str, Any]]] = []
    for idx, row in enumerate(reader, start=2):
        cleaned = {str(k).strip(): _clean(v) for k, v in row.items() if k is not None}
        if any(value for value in cleaned.values()):
            rows.append((idx, cleaned))
    return header, rows


def _read_xlsx_rows(file_bytes: bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    header = [_clean(value) for value in header_row]
    rows: list[tuple[int, dict[str, Any]]] = []
    for idx, raw in enumerate(rows_iter, start=2):
        values = [_clean(value) for value in raw]
        if not any(values):
            continue
        record = {
            header[col_idx]: values[col_idx]
            for col_idx in range(min(len(header), len(values)))
            if header[col_idx]
        }
        rows.append((idx, record))
    return header, rows


def _read_upload_rows(filename: str, file_bytes: bytes) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _read_csv_rows(file_bytes)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _read_xlsx_rows(file_bytes)
    raise BizError("仅支持 .xlsx / .csv 文件", code=40080)


def _header_error(header: list[str]) -> tuple[str | None, str | None]:
    normalized_header = [item.strip() for item in header if item.strip()]
    lower_header = [item.lower() for item in normalized_header]
    if "password" in lower_header:
        return "password", "模板不接受 password 列，初始密码必须由系统生成"
    if lower_header != TEMPLATE_COLUMNS:
        return "header", f"模板列必须为：{', '.join(TEMPLATE_COLUMNS)}"
    return None, None


def _normalize_row(
    raw: dict[str, Any],
    *,
    roles_by_code: dict[str, Role],
    existing_users: dict[str, User],
    seen_work_nos: set[str],
    actor_roles: list[str],
) -> tuple[str, str, str | None, dict[str, Any] | None]:
    work_no = _clean(raw.get("work_no"))
    display_name = _clean(raw.get("display_name"))
    email = _clean(raw.get("email"))
    role_code = normalize_role_code(_clean(raw.get("role_code")))
    scope_type = normalize_scope_type(_clean(raw.get("scope_type")) or SCOPE_GLOBAL)
    raw_scope_code = _clean(raw.get("scope_code"))
    is_active = _parse_bool(_clean(raw.get("is_active")))

    if not work_no:
        return ROW_SEVERITY_FATAL, "work_no", "工号必填", None
    if len(work_no) > 32:
        return ROW_SEVERITY_FATAL, "work_no", "工号长度不能超过 32", None
    if work_no in seen_work_nos:
        return ROW_SEVERITY_FATAL, "work_no", f"文件内重复工号：{work_no}", None
    seen_work_nos.add(work_no)

    if not display_name:
        return ROW_SEVERITY_FATAL, "display_name", "姓名必填", None
    if len(display_name) > 64:
        return ROW_SEVERITY_FATAL, "display_name", "姓名长度不能超过 64", None
    if email and (len(email) > 128 or not EMAIL_RE.match(email)):
        return ROW_SEVERITY_FATAL, "email", "邮箱格式错误", None
    if not role_code:
        return ROW_SEVERITY_FATAL, "role_code", "角色编码必填", None

    role = roles_by_code.get(role_code)
    if role is None or not role.is_active:
        return ROW_SEVERITY_FATAL, "role_code", f"角色不存在或已停用：{role_code}", None
    if scope_type not in VALID_SCOPE_TYPES:
        return ROW_SEVERITY_FATAL, "scope_type", "scope_type 仅支持 GLOBAL/CLASS/MAJOR/GRADE", None
    if scope_type == SCOPE_GLOBAL and raw_scope_code:
        return ROW_SEVERITY_FATAL, "scope_code", "GLOBAL 范围不应填写 scope_code", None
    if scope_type != SCOPE_GLOBAL and not raw_scope_code:
        return ROW_SEVERITY_FATAL, "scope_code", f"{scope_type} 范围必须填写 scope_code", None
    scope_code = build_scope_code(scope_type, raw_scope_code)
    if scope_code and len(scope_code) > 64:
        return ROW_SEVERITY_FATAL, "scope_code", "范围编码长度不能超过 64", None
    if is_active is None:
        return ROW_SEVERITY_FATAL, "is_active", "is_active 仅支持 true/false/启用/停用", None

    can_create, reason = _actor_can_create_role(actor_roles, role, scope_code)
    if not can_create:
        return ROW_SEVERITY_FATAL, "role_code", reason or "无权创建该角色", None

    normalized = {
        "work_no": work_no,
        "display_name": display_name,
        "email": email or None,
        "role_code": role_code,
        "scope_type": scope_type,
        "scope_code": scope_code,
        "is_active": is_active,
    }
    if work_no in existing_users:
        return (
            ROW_SEVERITY_WARN,
            "work_no",
            "工号已存在，提交时不会重置密码，仅补齐缺失角色/范围",
            normalized,
        )
    return ROW_SEVERITY_INFO, None, None, normalized


async def preview_import(
    db: AsyncSession,
    *,
    filename: str,
    file_bytes: bytes,
    mime_type: str | None,
    actor_user_id: int,
    actor_roles: list[str],
) -> AdminUserImportBatch:
    ensure_import_permission(actor_roles)
    if not file_bytes:
        raise BizError("空文件", code=40081)
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise BizError("文件超过 10MB 上限", code=40082)

    batch = await repo.create_batch(
        db,
        batch_no=_gen_batch_no(),
        filename=filename,
        file_size=len(file_bytes),
        mime_type=mime_type,
        operator_id=actor_user_id,
        operator_role=_actor_role_label(actor_roles),
    )

    total = ok = warn = fatal = 0
    try:
        header, upload_rows = _read_upload_rows(filename, file_bytes)
    except Exception as exc:
        await repo.add_batch_row(
            db,
            batch_id=batch.id,
            row_no=1,
            severity=ROW_SEVERITY_FATAL,
            result=ROW_RESULT_FAILED,
            field_name="file",
            message=f"文件解析失败：{exc}"[:512],
        )
        await repo.finalize_preview(
            db,
            batch,
            status=BATCH_STATUS_FAILED,
            total_rows=0,
            ok_rows=0,
            warn_rows=0,
            fatal_rows=1,
            summary={"header": [], "error": str(exc)[:256]},
        )
        await _audit_preview(db, batch, actor_user_id, actor_roles)
        await db.commit()
        await db.refresh(batch)
        return batch

    field, message = _header_error(header)
    if message:
        fatal = 1
        await repo.add_batch_row(
            db,
            batch_id=batch.id,
            row_no=1,
            severity=ROW_SEVERITY_FATAL,
            result=ROW_RESULT_FAILED,
            field_name=field,
            message=message,
            raw_data={"header": header},
        )

    role_codes = {
        normalize_role_code(_clean(raw.get("role_code")))
        for _, raw in upload_rows
        if normalize_role_code(_clean(raw.get("role_code")))
    }
    work_nos = {_clean(raw.get("work_no")) for _, raw in upload_rows if _clean(raw.get("work_no"))}
    roles_by_code = await repo.list_roles_by_codes(db, {code for code in role_codes if code})
    existing_users = await repo.list_users_by_work_no(db, work_nos)

    seen_work_nos: set[str] = set()
    if not message:
        for row_no, raw in upload_rows:
            total += 1
            severity, field_name, row_message, normalized = _normalize_row(
                raw,
                roles_by_code=roles_by_code,
                existing_users=existing_users,
                seen_work_nos=seen_work_nos,
                actor_roles=actor_roles,
            )
            if severity == ROW_SEVERITY_FATAL:
                fatal += 1
                result = ROW_RESULT_FAILED
            elif severity == ROW_SEVERITY_WARN:
                warn += 1
                ok += 1
                result = ROW_RESULT_VALID
            else:
                ok += 1
                result = ROW_RESULT_VALID
            await repo.add_batch_row(
                db,
                batch_id=batch.id,
                row_no=row_no,
                work_no=normalized["work_no"] if normalized else _clean(raw.get("work_no")) or None,
                role_code=normalized["role_code"] if normalized else normalize_role_code(_clean(raw.get("role_code"))),
                scope_code=normalized["scope_code"] if normalized else None,
                raw_data=raw,
                normalized_data=normalized,
                severity=severity,
                result=result,
                field_name=field_name,
                message=row_message,
            )

    status = BATCH_STATUS_VALIDATED if fatal == 0 else BATCH_STATUS_FAILED
    await repo.finalize_preview(
        db,
        batch,
        status=status,
        total_rows=total,
        ok_rows=ok,
        warn_rows=warn,
        fatal_rows=fatal,
        summary={"header": header, "template_columns": TEMPLATE_COLUMNS},
    )
    await _audit_preview(db, batch, actor_user_id, actor_roles)
    await db.commit()
    await db.refresh(batch)
    return batch


async def _audit_preview(
    db: AsyncSession,
    batch: AdminUserImportBatch,
    actor_user_id: int,
    actor_roles: list[str],
) -> None:
    await log_action(
        db,
        event_type="ADMIN_USER_IMPORT",
        entity_code="ADMIN_USER_IMPORT_BATCH",
        action="PREVIEW",
        entity_id=batch.id,
        actor_user_id=actor_user_id,
        actor_role=_actor_role_label(actor_roles),
        result_code="FAILED" if batch.fatal_rows else "SUCCESS",
        detail=build_audit_detail(
            metrics={
                "total": batch.total_rows,
                "ok": batch.ok_rows,
                "warn": batch.warn_rows,
                "fatal": batch.fatal_rows,
            },
            refs=[{"batch_no": batch.batch_no, "filename": batch.filename}],
        ),
        auto_flush=False,
    )


async def commit_import(
    db: AsyncSession,
    *,
    batch_id: int,
    actor_user_id: int,
    actor_roles: list[str],
    note: str | None = None,
) -> tuple[AdminUserImportBatch, list[AdminUserCredentialOut]]:
    ensure_import_permission(actor_roles)
    batch = await repo.get_batch(db, batch_id)
    if batch is None:
        raise NotFoundError("批次不存在")
    if batch.status != BATCH_STATUS_VALIDATED:
        raise BizError(f"批次状态 {batch.status} 不可提交，仅 VALIDATED 可提交", code=40083)
    if batch.fatal_rows > 0:
        raise BizError("批次存在致命错误，无法提交", code=40084)

    rows = await repo.list_batch_rows(db, batch.id)
    credentials: list[AdminUserCredentialOut] = []
    created = existing = role_granted = unchanged = 0
    try:
        for row in rows:
            if row.severity == ROW_SEVERITY_FATAL:
                continue
            normalized = row.normalized_data or {}
            role_code = normalized.get("role_code")
            scope_code = normalized.get("scope_code")
            role = await auth_repo.get_role_by_code(db, role_code)
            if role is None:
                raise BizError(f"角色不存在：{role_code}", code=40085)
            can_create, reason = _actor_can_create_role(actor_roles, role, scope_code)
            if not can_create:
                raise PermissionError(reason or "无权创建该角色")

            user = await auth_repo.get_user_by_work_no(db, normalized["work_no"])
            if user is None:
                initial_password = _generate_initial_password()
                user = User(
                    work_no=normalized["work_no"],
                    display_name=normalized["display_name"],
                    email=normalized.get("email"),
                    password_hash=hash_password(initial_password),
                    must_change_password=True,
                    is_active=bool(normalized.get("is_active", True)),
                )
                db.add(user)
                await db.flush()
                await auth_repo.ensure_user_role(
                    db,
                    user_id=user.id,
                    role_code=role_code,
                    scope_code=scope_code,
                    granted_by=actor_user_id,
                )
                credentials.append(
                    AdminUserCredentialOut(
                        work_no=user.work_no or "",
                        display_name=user.display_name,
                        role_code=role_code,
                        scope_code=scope_code,
                        initial_password=initial_password,
                    )
                )
                row.result = ROW_RESULT_CREATED
                row.message = "新账号已创建，初始密码仅在本次响应返回"
                created += 1
                await _audit_role_grant(
                    db,
                    user_id=user.id,
                    role_code=role_code,
                    scope_code=scope_code,
                    batch=batch,
                    actor_user_id=actor_user_id,
                    actor_roles=actor_roles,
                )
                continue

            existing += 1
            current_roles = await repo.list_roles_for_user(db, user.id)
            role_pair = (role_code, scope_code)
            existing_pairs = {(item.role_code, item.scope_code) for item in current_roles}
            if role_pair in existing_pairs:
                unchanged += 1
                row.result = ROW_RESULT_EXISTING
                row.message = "账号已存在且角色/范围已存在，未重置密码"
                continue

            await auth_repo.ensure_user_role(
                db,
                user_id=user.id,
                role_code=role_code,
                scope_code=scope_code,
                granted_by=actor_user_id,
            )
            role_granted += 1
            row.result = ROW_RESULT_ROLE_GRANTED
            row.message = "账号已存在，已补齐缺失角色/范围，未重置密码"
            await _audit_role_grant(
                db,
                user_id=user.id,
                role_code=role_code,
                scope_code=scope_code,
                batch=batch,
                actor_user_id=actor_user_id,
                actor_roles=actor_roles,
            )

        summary = {
            "created": created,
            "existing": existing,
            "role_granted": role_granted,
            "unchanged": unchanged,
            "note": note,
        }
        await repo.mark_committed(
            db,
            batch,
            created_rows=created,
            existing_rows=existing,
            role_granted_rows=role_granted,
            unchanged_rows=unchanged,
            summary={**(batch.summary or {}), "commit": summary},
        )
        await log_action(
            db,
            event_type="ADMIN_USER_IMPORT",
            entity_code="ADMIN_USER_IMPORT_BATCH",
            action="COMMIT",
            entity_id=batch.id,
            actor_user_id=actor_user_id,
            actor_role=_actor_role_label(actor_roles),
            detail=build_audit_detail(metrics=summary, refs=[{"batch_no": batch.batch_no}]),
            auto_flush=False,
        )
        await db.commit()
    except Exception:
        await db.rollback()
        raise

    await db.refresh(batch)
    return batch, credentials


async def _audit_role_grant(
    db: AsyncSession,
    *,
    user_id: int,
    role_code: str,
    scope_code: str | None,
    batch: AdminUserImportBatch,
    actor_user_id: int,
    actor_roles: list[str],
) -> None:
    await log_action(
        db,
        event_type="AUTH",
        entity_code="USER_ROLE",
        action="GRANT_ROLE",
        entity_id=user_id,
        actor_user_id=actor_user_id,
        actor_role=_actor_role_label(actor_roles),
        detail=build_audit_detail(
            changes=[{"role_code": role_code, "scope_code": scope_code}],
            refs=[{"batch_id": batch.id, "batch_no": batch.batch_no}],
        ),
        auto_flush=False,
    )


def build_template_file(fmt: str) -> tuple[bytes, str, str]:
    normalized = fmt.lower()
    if normalized == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(TEMPLATE_COLUMNS)
        writer.writerow(["T2026001", "张三", "zhangsan@example.edu.cn", "COUNSELOR", "GLOBAL", "", "true"])
        writer.writerow(["T2026002", "李四", "lisi@example.edu.cn", "CLASS_MONITOR", "CLASS", "CS2401", "true"])
        return (
            ("\ufeff" + buffer.getvalue()).encode("utf-8"),
            "admin-user-import-template.csv",
            "text/csv; charset=utf-8",
        )
    if normalized != "xlsx":
        raise BizError("format 仅支持 xlsx/csv", code=40086)
    wb = Workbook()
    ws = wb.active
    ws.title = "admin_users"
    ws.append(TEMPLATE_COLUMNS)
    ws.append(["T2026001", "张三", "zhangsan@example.edu.cn", "COUNSELOR", "GLOBAL", "", "true"])
    ws.append(["T2026002", "李四", "lisi@example.edu.cn", "CLASS_MONITOR", "CLASS", "CS2401", "true"])
    buf = io.BytesIO()
    wb.save(buf)
    return (
        buf.getvalue(),
        "admin-user-import-template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


async def build_error_report(db: AsyncSession, batch_id: int) -> tuple[bytes, str]:
    batch = await repo.get_batch(db, batch_id)
    if batch is None:
        raise NotFoundError("批次不存在")
    rows = await repo.list_batch_rows(db, batch_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    ws.append(
        [
            "row_no",
            "severity",
            "result",
            "field_name",
            "message",
            "work_no",
            "display_name",
            "email",
            "role_code",
            "scope_type",
            "scope_code",
            "is_active",
        ]
    )
    for row in rows:
        raw = row.raw_data or {}
        normalized = row.normalized_data or {}
        ws.append(
            [
                row.row_no,
                row.severity,
                row.result,
                row.field_name,
                row.message,
                row.work_no or raw.get("work_no"),
                normalized.get("display_name") or raw.get("display_name"),
                normalized.get("email") or raw.get("email"),
                row.role_code or raw.get("role_code"),
                normalized.get("scope_type") or raw.get("scope_type"),
                row.scope_code or raw.get("scope_code"),
                normalized.get("is_active") if normalized else raw.get("is_active"),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), f"admin-user-import-errors-{batch.batch_no}.xlsx"
